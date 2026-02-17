import csv
from uuid import uuid4
from math import isnan, isinf
import math
from django.db import connection, transaction, models
from rest_framework.exceptions import ValidationError
from psycopg2 import sql
from openpyxl import load_workbook
import logging

from src.core.bi_analysis.bi_datasets.models import DataSetField, DataSetTable, FileUpload
from src.core.bi_analysis.services.formula_to_sql import is_formula, formula_to_sql

logger = logging.getLogger(__name__)


class TableData:
    __slots__ = ('columns', 'rows')

    def __init__(self, columns, rows):
        self.columns = columns or []
        self.rows = rows or []

    def __len__(self):
        return len(self.rows)

    def limited(self, limit):
        if limit is None or limit <= 0 or len(self.rows) <= limit:
            return self
        return TableData(self.columns, self.rows[:limit])


def _normalize_header(header):
    if not header:
        return []
    normalized = []
    for col in header:
        if col is None:
            normalized.append('')
        else:
            normalized.append(str(col))
    return normalized


def _ensure_row(values, expected_len):
    row = list(values or [])
    if expected_len and len(row) < expected_len:
        row.extend([None] * (expected_len - len(row)))
    elif expected_len and len(row) > expected_len:
        row = row[:expected_len]
    return row


def _is_null(val):
    if val is None:
        return True
    if isinstance(val, float):
        return isnan(val)
    return False

def populate_initial_fields(dataset, temp_table_name, staging_table=None):
    """
    Создаёт DataSetField для каждой колонки temp_table_name с дефолтными type/aggregation.
    Если имя столбца вида <table>__<col> — ищет соответствующую DataSetTable.
    УСТАРЕВШАЯ ФУНКЦИЯ - используется только для обратной совместимости.
    """
    cols = introspect_columns(temp_table_name)
    ds_tables = {t.table_name: t for t in DataSetTable.objects.filter(dataset=dataset)}
    objs = []

    for idx, col in enumerate(cols):
        source_tbl = None
        if '__' in col:
            tbl_name, _ = col.split('__', 1)
            source_tbl = ds_tables.get(tbl_name)
        else:
            if staging_table:
                tbl_name = staging_table.table_name if hasattr(staging_table, 'table_name') else staging_table
                source_tbl = ds_tables.get(tbl_name)
            if not source_tbl:
                source_tbl = next(iter(ds_tables.values()), None)

        if not source_tbl:
            print(f"[WARNING] Не удалось найти source_table для колонки '{col}'")
            continue

        objs.append(DataSetField(
            dataset=dataset,
            name=col,
            source_table=source_tbl,
            source_column=col,
            order=idx
        ))
    DataSetField.objects.bulk_create(objs)


def populate_initial_fields_from_file(dataset, file_upload, source_table):
    """
    Создаёт DataSetField для каждой колонки из файла без создания таблицы в БД.
    Читает файл напрямую через pandas для получения списка колонок.
    """
    try:
        table = read_file_to_dataframe(
            file_upload.id,
            sheet_name=getattr(source_table, 'sheet_name', None),
            row_limit=0
        )
        cols = table.columns
    except Exception as e:
        # Fallback: используем columns_info если есть
        if file_upload.columns_info and 'columns' in file_upload.columns_info:
            cols = file_upload.columns_info['columns']
        else:
            raise ValidationError(f"Не удалось прочитать колонки из файла: {str(e)}")
    
    objs = []
    for idx, col in enumerate(cols):
        objs.append(DataSetField(
            dataset=dataset,
            name=col,
            source_table=source_table,
            source_column=col,
            order=idx
        ))
    
    if objs:
        DataSetField.objects.bulk_create(objs)

def create_temp_table_from_source(dataset):
    """
    УСТАРЕВШАЯ ФУНКЦИЯ - больше не создает temp таблицы.
    Оставлена для обратной совместимости.
    """
    print(f"[WARNING] create_temp_table_from_source вызвана - функция устарела")
    raise ValidationError("Создание temp таблиц больше не поддерживается. Используйте новую архитектуру с чтением файлов напрямую.")

def find_existing_temp_table_for_file(file_upload_id):
    """
    Ищет существующую staging таблицу для данного файла.
    Проверяет таблицы через DataSetTable и также все staging_ таблицы с нужными колонками.
    """
    try:
        try:
            upload = FileUpload.objects.get(pk=file_upload_id)
        except FileUpload.DoesNotExist:
            logger.error(f"FileUpload с id={file_upload_id} не найден")
            return None
        
        # Сначала проверяем через DataSetTable
        existing_tables = DataSetTable.objects.filter(file_upload_id=file_upload_id)
        for ds_table in existing_tables:
            # Ищем staging или temp таблицы (temp для обратной совместимости)
            if ds_table.table_name.startswith(('staging_', 'temp_')):
                if table_exists(ds_table.table_name):
                    return ds_table.table_name
        
        # Если не нашли через DataSetTable, проверяем все staging_ таблицы
        # Загружаем ожидаемые колонки из файла для сравнения
        cols_from_file = []
        if upload.file_type in ('xlsx', 'csv', 'txt'):
            table = read_file_to_dataframe(upload.id, row_limit=0)
            cols_from_file = table.columns
        else:
            cols_from_file = []

        if cols_from_file:
            # Ищем все staging_ таблицы с таким же набором колонок
            # Также проверяем temp_ для обратной совместимости
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT tablename 
                    FROM pg_tables 
                    WHERE schemaname = 'public' 
                    AND (tablename LIKE 'staging_%' OR tablename LIKE 'temp_%')
                """)
                staging_tables = [row[0] for row in cursor.fetchall()]
                
                for staging_table in staging_tables:
                    try:
                        existing_cols = introspect_columns(staging_table)
                        # Сравниваем наборы колонок
                        if set(existing_cols) == set(cols_from_file):
                            return staging_table
                    except:
                        continue
    except Exception as e:
        print(f"[FIND STAGING TABLE] Error: {e}")
    
    return None

def import_file_upload_to_table(file_upload_id, dataset=None, reuse_existing=True):
    """
    УСТАРЕВШАЯ ФУНКЦИЯ - больше не создает таблицы в БД.
    Оставлена для обратной совместимости со старым кодом.
    
    НОВАЯ АРХИТЕКТУРА: данные читаются напрямую из файлов через pandas
    без создания материализованных таблиц.
    
    Эта функция больше не должна вызываться в новом коде.
    Используйте read_file_to_dataframe() вместо этого.
    """
    # Не создаем таблицы - это устаревший подход
    # Возвращаем None или пустую строку для обратной совместимости
    print(f"[WARNING] import_file_upload_to_table вызвана для file_id={file_upload_id} - функция устарела")
    return None


def _create_table_and_load(staging, cols, rows):
    """
    Общая логика: создаём staging-таблицу TEXT и массово вставляем rows.
    """
    col_defs = ", ".join(f'"{c}" TEXT' for c in cols)

    placeholders = ", ".join(["%s"] * len(cols))
    insert_sql   = f'''
        INSERT INTO "{staging}" ({", ".join(f'"{c}"' for c in cols)})
        VALUES ({placeholders})
    '''

    with connection.cursor() as cursor:
        cursor.execute(f'CREATE TABLE "{staging}" ({col_defs});')
        cursor.executemany(insert_sql, rows)

    return staging

def introspect_columns(temp_table_name):
    """
    Возвращает список (column_name, data_type) для временной таблицы.
    """
    safe_name = temp_table_name.replace('"', '""')
    sql = f'SELECT * FROM "{safe_name}" LIMIT 0;'
    with connection.cursor() as cursor:
        cursor.execute(sql)
        return [col[0] for col in cursor.description]
    
def table_exists(table_name):
    with connection.cursor() as cursor:
        cursor.execute("SELECT to_regclass(%s)", [table_name])
        exists = cursor.fetchone()[0] is not None
        print(f"[TABLE EXISTS] {table_name}: {exists}")
        return exists
    
def safe_drop_table(table_name):
    """
    Безопасно удаляет таблицу, если она есть.
    """
    with connection.cursor() as cursor:
        cursor.execute(f'DROP TABLE IF EXISTS "{table_name}" CASCADE;')

def auto_join_table(dataset, table, left_column, right_column,
                    join_type: str = "INNER JOIN"):
    """
    Добавляет JOIN к датасету. Теперь не создает материализованные таблицы,
    а только обновляет метаданные DataSetTable.
    """
    if "JOIN" not in join_type.upper():
        join_type = f"{join_type.strip().upper()} JOIN"

    # Получаем главную таблицу для проверки существования
    main_table = dataset.tables.filter(joined_on_type__isnull=True).first()
    if not main_table:
        raise ValueError("Не найдена главная таблица для датасета")
    
    main_table_name = main_table.table_name
    if not table_exists(main_table_name):
        raise ValueError(f"Таблица {main_table_name} не найдена, JOIN невозможен")
    
    if not table_exists(table.table_name):
        raise ValueError(f"Таблица {table.table_name} не найдена, JOIN невозможен")

    # Проверяем наличие общих значений для валидации JOIN
    with connection.cursor() as c:
        c.execute(f'SELECT DISTINCT "{left_column}" FROM "{main_table_name}" LIMIT 5000')
        main_vals = {r[0] for r in c.fetchall()}
        c.execute(f'SELECT DISTINCT "{right_column}" FROM "{table.table_name}" LIMIT 5000')
        join_vals = {r[0] for r in c.fetchall()}

    if not (main_vals & join_vals):
        raise ValueError("Нет общих значений; авто-JOIN прерван")

    # Обновляем метаданные таблицы для JOIN
    # joined_on_left - колонка в главной (или предыдущей) таблице
    # joined_on_right - колонка в присоединяемой таблице
    table.joined_on_type = join_type
    table.joined_on_left = left_column
    table.joined_on_right = right_column
    table.save(update_fields=['joined_on_type', 'joined_on_left', 'joined_on_right'])

    # Синхронизируем поля датасета после добавления JOIN
    # (можно добавить новые поля из присоединенной таблицы)
    sync_dataset_fields_after_all_joins(dataset)

    return left_column

def rebuild_dataset_joins(dataset):
    """
    Перестраивает JOIN'ы для датасета на основе метаданных.
    Теперь не создает материализованные таблицы, только обновляет метаданные.
    """
    base_tbl = dataset.tables.filter(joined_on_type__isnull=True).first()
    if not base_tbl:
        raise ValueError("Не найдена главная таблица")

    # Для файловых источников таблицы не нужны - данные читаются напрямую из файлов
    # Для БД источников проверяем существование таблицы только если это реальная таблица БД
    if base_tbl.file_upload_id is None and base_tbl.table_name:
        # Это БД таблица, проверяем существование
        if not table_exists(base_tbl.table_name):
            raise ValueError(f"Таблица {base_tbl.table_name} не существует")

    # Обновляем table_ref на главную таблицу (для обратной совместимости)
    dataset.table_ref = base_tbl.table_name
    dataset.save(update_fields=["table_ref"])

    # Обновляем метаданные JOIN'ов для всех присоединенных таблиц
    for t in (dataset.tables.filter(joined_on_type__isnull=False).order_by("id")):
        if not t.joined_on_left or not t.joined_on_right:
            continue
            
        # Для файловых источников таблицы не нужны - данные читаются напрямую из файлов
        # Для БД источников проверяем существование таблицы
        if t.file_upload_id is None and t.table_name:
            # Это БД таблица, проверяем существование
            if not table_exists(t.table_name):
                print(f"[WARNING] Таблица {t.table_name} для JOIN не существует, пропускаем")
                continue
        
        # Проверяем валидность JOIN (есть ли общие значения)
        # Только для БД источников (для файловых источников проверка не нужна, данные читаются на лету)
        if base_tbl.file_upload_id is None and t.file_upload_id is None:
            # Обе таблицы - БД источники, проверяем через SQL
            try:
                main_table_name = base_tbl.table_name
                with connection.cursor() as c:
                    c.execute(f'SELECT DISTINCT "{t.joined_on_left}" FROM "{main_table_name}" LIMIT 5000')
                    main_vals = {r[0] for r in c.fetchall()}
                    c.execute(f'SELECT DISTINCT "{t.joined_on_right}" FROM "{t.table_name}" LIMIT 5000')
                    join_vals = {r[0] for r in c.fetchall()}
                
                if not (main_vals & join_vals):
                    print(f"[WARNING] Нет общих значений для JOIN таблицы {t.id}, пропускаем")
                    continue
            except Exception as e:
                print(f"[WARNING] Ошибка проверки JOIN для таблицы {t.id}: {e}")
                continue
        # Для файловых источников проверку пропускаем - валидация будет при выполнении запроса
        
        print(f"Table id={t.id} joined_on_type={t.joined_on_type} joined_on_left={t.joined_on_left} joined_on_right={t.joined_on_right}")
    
    # Синхронизируем поля после всех JOIN'ов
    sync_dataset_fields_after_all_joins(dataset)


def sync_dataset_fields_after_all_joins(dataset):
    """
    Синхронизирует поля датасета после всех JOIN'ов.
    Строит запрос и проверяет доступные колонки.
    """
    try:
        query, display_columns = build_dataset_query(dataset, limit=0)

        with connection.cursor() as cursor:
            cursor.execute(query)
            columns = display_columns if display_columns else [col[0] for col in cursor.description]
        
        # Получаем существующие поля
        existing_fields = {f.name: f for f in DataSetField.objects.filter(dataset=dataset)}
        
        # Добавляем недостающие поля (используя логику определения source_table)
        for idx, col_name in enumerate(columns):
            if col_name not in existing_fields:
                # Определяем source_table для нового поля
                # Пробуем найти колонку в одной из таблиц датасета
                source_table = None
                for table in dataset.tables.all():
                    try:
                        # Для файловых источников используем columns_info
                        if table.file_upload_id:
                            if table.columns_info:
                                table_cols = table.columns_info
                            else:
                                # Если columns_info нет, пробуем получить из FileUpload
                                from src.core.bi_analysis.bi_datasets.models import FileUpload
                                try:
                                    file_upload = FileUpload.objects.get(pk=table.file_upload_id)
                                    table_cols = file_upload.columns_info or []
                                except:
                                    continue
                        else:
                            # Для БД источников используем introspect_columns
                            table_cols = introspect_columns(table.table_name)
                        
                        # Проверяем, есть ли колонка в таблице
                        if isinstance(table_cols, list):
                            # Если это список строк
                            if col_name in table_cols:
                                source_table = table
                                break
                        elif isinstance(table_cols, dict):
                            # Если это словарь с информацией о колонках
                            if col_name in table_cols or any(col.get('name') == col_name for col in table_cols.values() if isinstance(col, dict)):
                                source_table = table
                                break
                    except Exception as e:
                        # Пропускаем ошибки при проверке таблицы
                        continue
                
                if source_table:
                    DataSetField.objects.create(
                        dataset=dataset,
                        name=col_name,
                        source_table=source_table,
                        source_column=col_name,
                        order=idx
                    )
    except Exception as e:
        print(f"[SYNC FIELDS AFTER JOINS] Ошибка: {e}")
        # Не прерываем выполнение

def create_temp_table_from_staging(staging_name):
    """
    УСТАРЕВШАЯ ФУНКЦИЯ - больше не создает temp таблицы.
    Оставлена для обратной совместимости.
    """
    print(f"[WARNING] create_temp_table_from_staging вызвана для {staging_name} - функция устарела")
    raise ValidationError("Создание temp таблиц больше не поддерживается. Используйте новую архитектуру с чтением файлов напрямую.")

def get_columns_with_aliases(left_table, right_table):
    left_cols = introspect_columns(left_table)
    right_cols = introspect_columns(right_table)
    left_set = set(left_cols)
    columns = []

    columns += [f'a."{col}" AS "{col}"' for col in left_cols]
    for col in right_cols:
        if col in left_set:
            columns.append(f'b."{col}" AS "{col}__right"')
        else:
            columns.append(f'b."{col}" AS "{col}"')
    return columns

def ensure_temp_table_exists(ds_table):
    """
    УСТАРЕВШАЯ ФУНКЦИЯ - больше не создает таблицы.
    Оставлена для обратной совместимости, но не выполняет никаких действий.
    Для файловых источников таблицы не нужны - данные читаются напрямую из файлов.
    """
    # Функция больше не создает таблицы - они не нужны при новой архитектуре
    # Для файловых источников данные читаются напрямую через pandas
    # Для БД источников таблицы уже существуют в БД
    pass
    
def sync_dataset_fields_with_current_table(dataset):
    table_name = dataset.table_ref
    columns = introspect_columns(table_name)
    existing_fields = {f.source_column: f for f in DataSetField.objects.filter(dataset=dataset)}
    ds_tables = {t.table_name: t for t in DataSetTable.objects.filter(dataset=dataset)}

    # Добавить/обновить поля
    for idx, col in enumerate(columns):
        if col in existing_fields:
            field = existing_fields[col]
            # Можно обновить order, тип данных и т.д., если требуется
            field.order = idx
            field.save(update_fields=['order'])
        else:
            # Определи source_table (по логике как раньше)
            source_tbl = None
            for t in ds_tables.values():
                if col in introspect_columns(t.table_name):
                    source_tbl = t
                    break
            DataSetField.objects.create(
                dataset=dataset,
                name=col,
                source_table=source_tbl,
                source_column=col,
                order=idx
            )

    # Удалить устаревшие поля
    for col, field in existing_fields.items():
        if col not in columns:
            field.delete()


def _auto_bind_file_to_main_table(dataset, main_table):
    """
    Пытается автоматически привязать файловый источник к главной таблице датасета,
    если file_upload_id потерян (например, после замены/удаления файла) и в подключении
    остался ровно один файл.
    """
    if getattr(main_table, "file_upload_id", None) is not None:
        return False

    try:
        table_name = getattr(main_table, "table_name", "") or ""
        looks_like_file = isinstance(table_name, str) and table_name.lower().endswith(
            (".xlsx", ".xls", ".csv", ".txt", ".bin")
        )
        connection_id = getattr(main_table, "connection_id", None)
        if not looks_like_file or not connection_id:
            return False

        candidates = FileUpload.objects.filter(connection_id=connection_id)
        if candidates.count() != 1:
            return False

        new_upload = candidates.first()
        main_table.file_upload = new_upload
        if hasattr(main_table, "display_name"):
            main_table.display_name = new_upload.original_filename
        if hasattr(main_table, "columns_info"):
            main_table.columns_info = new_upload.columns_info

        update_fields = ["file_upload"]
        if hasattr(main_table, "display_name"):
            update_fields.append("display_name")
        if hasattr(main_table, "columns_info"):
            update_fields.append("columns_info")
        main_table.save(update_fields=update_fields)

        # Если у датасета только одна таблица, пересобираем поля под новый файл
        try:
            if hasattr(dataset, "tables") and dataset.tables.count() == 1:
                DataSetField.objects.filter(dataset=dataset).delete()
                populate_initial_fields_from_file(dataset, new_upload, main_table)
        except Exception as e:
            logger.warning(f"_auto_bind_file_to_main_table: failed to rebuild fields: {e}")

        return True
    except Exception as e:
        logger.warning(f"_auto_bind_file_to_main_table failed: {e}")
        return False


def _read_excel_table(path, sheet_name=None, row_limit=None):
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise ValidationError(f"Лист '{sheet_name}' не найден в файле")
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]
    except IndexError:
        wb.close()
        return TableData([], [])

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    columns = _normalize_header(header)

    if row_limit == 0:
        wb.close()
        return TableData(columns, [])

    rows = []
    for row in rows_iter:
        rows.append(_ensure_row(row, len(columns)))
        if row_limit and row_limit > 0 and len(rows) >= row_limit:
            break

    wb.close()
    return TableData(columns, rows)


def _read_csv_table(path, row_limit=None, encoding='cp1251'):
    with open(path, 'r', encoding=encoding, errors='replace', newline='') as f:
        reader = csv.reader(f)
        header = next(reader, None)
        columns = _normalize_header(header)

        if row_limit == 0:
            return TableData(columns, [])

        rows = []
        for row in reader:
            rows.append(_ensure_row(row, len(columns)))
            if row_limit and row_limit > 0 and len(rows) >= row_limit:
                break

    return TableData(columns, rows)


def count_file_rows(file_upload_id, sheet_name=None):
    """
    Подсчитывает общее количество строк в файле без загрузки всех данных.
    Использует polars для эффективного подсчёта.
    """
    try:
        upload = FileUpload.objects.get(pk=file_upload_id)
    except FileUpload.DoesNotExist:
        raise ValidationError(f"FileUpload с id={file_upload_id} не найден")
    
    if not upload.file:
        raise ValidationError(f"FileUpload {file_upload_id} не имеет файла")
    
    path = upload.file.path
    
    # Проверяем, является ли файл бинарным
    from src.core.bi_analysis.bi_datasets.binary_storage import is_binary_file, read_from_binary
    
    if is_binary_file(path) or upload.file_type == 'bin':
        # Для бинарных файлов читаем все данные для подсчёта
        try:
            columns, rows = read_from_binary(path, row_limit=None)
            return len(rows)
        except Exception as e:
            logger.error(f"Ошибка чтения бинарного файла для подсчёта: {str(e)}")
            raise ValidationError(f"Ошибка чтения бинарного файла: {str(e)}")
    
    # Используем polars для эффективного подсчёта
    try:
        import polars as pl
        sheet = sheet_name or getattr(upload, 'sheet_name', None)
        
        if upload.file_type == 'xlsx':
            if sheet:
                df = pl.read_excel(path, sheet_name=sheet)
            else:
                df = pl.read_excel(path)
        elif upload.file_type in ('csv', 'txt'):
            df = pl.read_csv(path)
        else:
            raise ValidationError(f"Неподдерживаемый тип файла: {upload.file_type}")
        
        return len(df)
    except ImportError:
        # Если polars недоступен, используем fallback
        logger.warning("Polars недоступен для подсчёта строк, используется fallback метод")
        # Читаем файл с большим лимитом для подсчёта
        table_data = read_file_to_dataframe(file_upload_id, sheet_name, row_limit=1000000)
        return len(table_data.rows)
    except Exception as e:
        logger.warning(f"Ошибка подсчёта строк через polars: {str(e)}, используется fallback")
        # Fallback: читаем файл для подсчёта
        table_data = read_file_to_dataframe(file_upload_id, sheet_name, row_limit=1000000)
        return len(table_data.rows)


def read_file_to_dataframe(file_upload_id, sheet_name=None, row_limit=None, use_polars=True):
    """
    Читает файл в табличную структуру.
    Использует polars для максимальной производительности, если доступен.
    Поддерживает чтение из бинарных файлов .bin.
    """
    try:
        upload = FileUpload.objects.get(pk=file_upload_id)
    except FileUpload.DoesNotExist:
        raise ValidationError(f"FileUpload с id={file_upload_id} не найден")
    
    if not upload.file:
        raise ValidationError(f"FileUpload {file_upload_id} не имеет файла")
    
    path = upload.file.path
    
    # Проверяем, является ли файл бинарным
    from src.core.bi_analysis.bi_datasets.binary_storage import is_binary_file, read_from_binary
    
    # Если row_limit=None, читаем все данные. Если указан, используем его.
    # Это универсальное решение для CSV и Excel через polars
    
    if is_binary_file(path) or upload.file_type == 'bin':
        # Читаем из бинарного файла
        try:
            # Если row_limit=None, читаем все данные
            columns, rows = read_from_binary(path, row_limit=row_limit)
            return TableData(columns, rows)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Ошибка чтения бинарного файла: {str(e)}")
            raise ValidationError(f"Ошибка чтения бинарного файла: {str(e)}")
    
    # Используем polars для чтения файлов (универсальное решение для CSV и Excel)
    if use_polars:
        try:
            from src.core.bi_analysis.tasks import _read_file_with_polars
            # Передаем row_limit как есть: None = читать все данные, число = лимит
            columns, rows = _read_file_with_polars(file_upload_id, sheet_name, row_limit)
            return TableData(columns, rows)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Не удалось использовать polars, используется fallback метод: {str(e)}")
    
    # Fallback на старый метод (только если polars недоступен)
    sheet = sheet_name or getattr(upload, 'sheet_name', None)

    if upload.file_type == 'xlsx':
        return _read_excel_table(path, sheet, row_limit=row_limit)
    elif upload.file_type in ('csv', 'txt'):
        return _read_csv_table(path, row_limit=row_limit)
    else:
        raise ValidationError(f"Неподдерживаемый тип файла: {upload.file_type}")


def dataframe_to_sql_values(table, table_alias='t0', row_limit=None):
    """
    Преобразует TableData в SQL (VALUES ...).
    
    Args:
        table: TableData объект с данными
        table_alias: алиас таблицы в SQL запросе
        row_limit: лимит строк для включения в VALUES. Если None, ограничиваем до разумного максимума.
    """
    if not table.columns:
        return sql.SQL('(SELECT NULL::text AS col WHERE FALSE)'), None

    # Лимиты полностью убраны - всегда используем все доступные данные
    limited = table
    
    if len(limited) == 0:
        return sql.SQL('(SELECT NULL::text AS col WHERE FALSE)'), None

    def sanitize(val):
        if _is_null(val):
            return 'NULL'
        # Обрабатываем разные типы данных
        if isinstance(val, bool):
            return 'TRUE' if val else 'FALSE'
        if isinstance(val, (int, float)):
            # Проверяем на infinity
            if isinstance(val, float) and (math.isinf(val) or math.isnan(val)):
                return 'NULL'
            return str(val)
        # Для строк экранируем кавычки
        val_str = str(val).replace("'", "''")
        return f"'{val_str}'"

    data_rows = []
    for row in limited.rows:
        # Проверяем, что строка не пустая
        if not row:
            continue
        sanitized = [sanitize(val) for val in row]
        # Проверяем, что после санитизации есть данные
        if sanitized:
            data_rows.append(f"({', '.join(sanitized)})")
    
    # Если после обработки нет строк, возвращаем пустой результат
    if not data_rows:
        return sql.SQL('(SELECT NULL::text AS col WHERE FALSE)'), None

    # Используем короткие уникальные имена (col_0, col_1, ...), т.к. в PostgreSQL
    # идентификаторы обрезаются до 63 байт; длинные кириллические имена дают дубликаты
    col_defs = [f'"col_{i}"' for i in range(len(limited.columns))]

    values_clause = ',\n'.join(data_rows)
    col_list = ', '.join(col_defs)

    values_query = f"""
        (VALUES {values_clause}) AS {table_alias}({col_list})
    """

    return sql.SQL(values_query), None


def build_dataset_query(dataset, select_fields=None, limit=None, offset=None, where_clause=None, search=None):
    """
    Строит SQL-запрос для датасета на основе метаданных таблиц и полей.
    Для файловых источников читает данные напрямую из файлов без создания таблиц.
    
    Args:
        dataset: объект Dataset
        select_fields: список имен полей для SELECT (None = все поля)
        limit: лимит строк
        offset: смещение для пагинации
        where_clause: дополнительное условие WHERE
        search: строка поиска (будет добавлена в WHERE как LIKE по всем текстовым полям)
    
    Returns:
        SQL объект для выполнения запроса
    """
    # Получаем главную таблицу (без JOIN)
    main_table = dataset.tables.filter(joined_on_type__isnull=True).first()
    if not main_table:
        raise ValueError("Не найдена главная таблица для датасета")
    
    # Алиас для главной таблицы
    main_alias = 't0'
    
    # Определяем, является ли источник файловым
    is_file_source = main_table.file_upload_id is not None

    # Пытаемся автоматически восстановить файловый источник, если он потерян
    if not is_file_source:
        if _auto_bind_file_to_main_table(dataset, main_table):
            is_file_source = True

    main_table_columns = None
    join_table_columns = {}

    if is_file_source:
        # Для файловых источников читаем данные напрямую из файла
        # Вычисляем сколько строк нужно прочитать: limit + offset (если есть)
        # Для больших файлов читаем все данные, но лимит будет применен в SQL
        file_read_limit = None
        if limit is not None:
            # Читаем достаточно данных с учетом offset
            file_read_limit = limit + (offset or 0)
            # Но если лимит очень большой, читаем все данные для эффективности
            if file_read_limit > 50000:
                file_read_limit = None
        
        try:
            df = read_file_to_dataframe(
                main_table.file_upload_id,
                sheet_name=getattr(main_table, 'sheet_name', None),
                row_limit=file_read_limit
            )
            main_table_columns = list(df.columns)
            # Передаем None для row_limit в dataframe_to_sql_values, чтобы включить все прочитанные данные
            # Лимит и offset будут применены в SQL запросе через LIMIT/OFFSET clauses
            from_with_alias, _ = dataframe_to_sql_values(df, main_alias, row_limit=None)
        except Exception as e:
            raise ValueError(f"Ошибка чтения файла: {str(e)}")
    else:
        # Для БД источников используем прямое подключение к таблице
        table_name = main_table.table_name
        if '.' in table_name:
            schema, table = table_name.split('.', 1)
            from_clause = sql.SQL('{}.{}').format(
                sql.Identifier(schema),
                sql.Identifier(table)
            )
        else:
            from_clause = sql.Identifier(table_name)
        
        from_with_alias = sql.SQL('{} AS {}').format(from_clause, sql.Identifier(main_alias))
    
    # Собираем JOIN'ы
    joins = []
    joined_tables = dataset.tables.filter(joined_on_type__isnull=False).order_by('id')
    
    # Маппинг для отслеживания алиасов таблиц
    table_id_to_alias = {main_table.id: main_alias}
    
    for idx, join_table in enumerate(joined_tables, start=1):
        if not join_table.joined_on_left or not join_table.joined_on_right:
            continue
        
        # Алиас для таблицы в JOIN
        table_alias = f't{idx}'
        table_id_to_alias[join_table.id] = table_alias
        
        # Определяем, является ли присоединяемая таблица файловым источником
        is_join_file_source = join_table.file_upload_id is not None
        
        if is_join_file_source:
            # Для файловых источников читаем данные напрямую из файла
            # Для JOIN таблиц также читаем все данные, лимит будет применен к итоговому запросу
            try:
                df_join = read_file_to_dataframe(
                    join_table.file_upload_id,
                    sheet_name=getattr(join_table, 'sheet_name', None),
                    row_limit=None  # Читаем все данные для JOIN
                )
                join_table_columns[join_table.id] = list(df_join.columns)
                # Передаем None для row_limit, чтобы включить все данные для JOIN
                join_table_ref, _ = dataframe_to_sql_values(df_join, table_alias, row_limit=None)
            except Exception as e:
                raise ValueError(f"Ошибка чтения файла для JOIN: {str(e)}")
        else:
            # Для БД источников используем прямое подключение к таблице
            join_table_name = join_table.table_name
            if '.' in join_table_name:
                join_schema, join_table_only = join_table_name.split('.', 1)
                join_table_ref = sql.SQL('{}.{}').format(
                    sql.Identifier(join_schema),
                    sql.Identifier(join_table_only)
                )
            else:
                join_table_ref = sql.Identifier(join_table_name)
            
            join_table_ref = sql.SQL('{} AS {}').format(join_table_ref, sql.Identifier(table_alias))
        
        # Тип JOIN
        join_type = (join_table.joined_on_type or 'LEFT JOIN').strip().upper()
        if 'JOIN' not in join_type:
            join_type = f'{join_type} JOIN'
        
        # Определяем левую таблицу для JOIN
        # joined_on_left - это колонка в главной таблице (или предыдущей joined таблице)
        left_table_alias = main_alias  # По умолчанию главная таблица
        
        # Для файловых источников подзапрос отдаёт col_0, col_1, ... — используем индекс
        left_col_ref = (
            f'col_{main_table_columns.index(join_table.joined_on_left)}'
            if is_file_source and main_table_columns and join_table.joined_on_left in main_table_columns
            else join_table.joined_on_left
        )
        right_col_ref = (
            f'col_{join_table_columns[join_table.id].index(join_table.joined_on_right)}'
            if is_join_file_source and join_table.joined_on_right in join_table_columns.get(join_table.id, [])
            else join_table.joined_on_right
        )
        join_condition = sql.SQL('{}.{} = {}.{}').format(
            sql.Identifier(left_table_alias),
            sql.Identifier(left_col_ref),
            sql.Identifier(table_alias),
            sql.Identifier(right_col_ref)
        )
        
        if is_join_file_source:
            # Для файловых источников VALUES уже содержит алиас, поэтому просто добавляем JOIN
            join_clause = sql.SQL('{} {} ON {}').format(
                sql.SQL(join_type),
                join_table_ref,
                join_condition
            )
        else:
            join_clause = sql.SQL('{} {} ON {}').format(
                sql.SQL(join_type),
                join_table_ref,
                join_condition
            )
        joins.append(join_clause)
    
    # Строим SELECT часть
    if select_fields is None:
        # Берем все поля датасета
        fields = dataset.fields.all().order_by('order')
    else:
        # Берем только указанные поля
        fields = dataset.fields.filter(name__in=select_fields).order_by('order')
    
    select_parts = []
    table_aliases = table_id_to_alias.copy()
    display_columns = []

    def _col_ref(table_id, source_column, fallback_out_idx=None):
        if table_id == main_table.id and main_table_columns:
            if source_column in main_table_columns:
                return f'col_{main_table_columns.index(source_column)}'
            if is_file_source and fallback_out_idx is not None and fallback_out_idx < len(main_table_columns):
                return f'col_{fallback_out_idx}'
        if table_id in join_table_columns and source_column in join_table_columns[table_id]:
            return f'col_{join_table_columns[table_id].index(source_column)}'
        return source_column

    field_refs = {}
    for f in dataset.fields.all().order_by('order'):
        table_alias = table_aliases.get(f.source_table.id, main_alias)
        col_ref = _col_ref(f.source_table.id, f.source_column)
        field_refs[f.name] = sql.SQL('{}.{}').format(
            sql.Identifier(table_alias),
            sql.Identifier(col_ref)
        )

    for out_idx, field in enumerate(fields):
        if field.expression:
            if is_formula(field.expression):
                field_expr, err = formula_to_sql(field.expression, field_refs, {})
                if err:
                    raise ValueError(f"Ошибка в формуле поля {field.name!r}: {err}")
                if field_expr is None:
                    raise ValueError(f"Ошибка в формуле поля {field.name!r}")
            else:
                field_expr = sql.SQL(field.expression)
        else:
            table_alias = table_aliases.get(field.source_table.id, main_alias)
            col_ref = _col_ref(field.source_table.id, field.source_column, fallback_out_idx=out_idx)
            field_expr = sql.SQL('{}.{}').format(
                sql.Identifier(table_alias),
                sql.Identifier(col_ref)
            )
        select_parts.append(
            sql.SQL('{} AS {}').format(
                field_expr,
                sql.Identifier(f'out_{out_idx}')
            )
        )
        display_columns.append(field.name)

    if not select_parts:
        if is_file_source and main_table_columns:
            for out_idx, col in enumerate(main_table_columns):
                select_parts.append(
                    sql.SQL('{}.{} AS {}').format(
                        sql.Identifier(main_alias),
                        sql.Identifier(f'col_{out_idx}'),
                        sql.Identifier(f'out_{out_idx}')
                    )
                )
            display_columns = main_table_columns
        else:
            select_parts.append(sql.SQL('{}.*').format(sql.Identifier(main_alias)))
    
    # Собираем итоговый запрос
    query_parts = [
        sql.SQL('SELECT {}').format(sql.SQL(', ').join(select_parts)),
        sql.SQL('FROM {}').format(from_with_alias)
    ]
    
    # Добавляем JOIN'ы
    query_parts.extend(joins)
    
    # Собираем условия WHERE
    where_conditions = []
    
    if where_clause:
        where_conditions.append(sql.SQL(where_clause))
    
    # Добавляем поиск по всем текстовым полям
    if search:
        search_conditions = []
        search_fields = dataset.fields.all() if select_fields is None else dataset.fields.filter(name__in=select_fields)
        for field in search_fields:
            table_alias = table_aliases.get(field.source_table.id, main_alias)
            col_ref = _col_ref(field.source_table.id, field.source_column)
            search_conditions.append(
                sql.SQL('CAST({}.{} AS TEXT) ILIKE {}').format(
                    sql.Identifier(table_alias),
                    sql.Identifier(col_ref),
                    sql.Literal(f'%{search}%')
                )
            )
        if search_conditions:
            where_conditions.append(sql.SQL('({})').format(sql.SQL(' OR ').join(search_conditions)))
    
    # Добавляем WHERE если есть условия
    if where_conditions:
        query_parts.append(sql.SQL('WHERE {}').format(sql.SQL(' AND ').join(where_conditions)))
    
    # Добавляем OFFSET если есть
    if offset is not None and offset > 0:
        query_parts.append(sql.SQL('OFFSET {}').format(sql.Literal(offset)))
    
    # Добавляем LIMIT если есть
    if limit is not None:
        query_parts.append(sql.SQL('LIMIT {}').format(sql.Literal(limit)))
    
    final_query = sql.SQL(' ').join(query_parts)
    return final_query, display_columns if display_columns else None


def build_dataset_count_query(dataset, where_clause=None, search=None):
    """
    Строит SQL-запрос COUNT(*) для датасета.
    Используется для подсчёта общего количества строк без загрузки всех данных.
    
    Args:
        dataset: объект Dataset
        where_clause: дополнительное условие WHERE
        search: строка поиска (будет добавлена в WHERE как LIKE по всем текстовым полям)
    
    Returns:
        SQL объект для выполнения запроса COUNT(*)
    """
    # Получаем главную таблицу (без JOIN)
    main_table = dataset.tables.filter(joined_on_type__isnull=True).first()
    if not main_table:
        raise ValueError("Не найдена главная таблица для датасета")
    
    # Алиас для главной таблицы
    main_alias = 't0'
    
    # Определяем, является ли источник файловым
    is_file_source = main_table.file_upload_id is not None

    # Пытаемся автоматически восстановить файловый источник, если он потерян
    if not is_file_source:
        if _auto_bind_file_to_main_table(dataset, main_table):
            is_file_source = True
    
    main_table_columns = None
    join_table_columns = {}
    
    if is_file_source:
        # Для файловых источников читаем данные напрямую из файла
        try:
            df = read_file_to_dataframe(
                main_table.file_upload_id,
                sheet_name=getattr(main_table, 'sheet_name', None),
                row_limit=None
            )
            main_table_columns = list(df.columns)
            from_with_alias, _ = dataframe_to_sql_values(df, main_alias, row_limit=None)
        except Exception as e:
            raise ValueError(f"Ошибка чтения файла: {str(e)}")
    else:
        table_name = main_table.table_name
        if '.' in table_name:
            schema, table = table_name.split('.', 1)
            from_clause = sql.SQL('{}.{}').format(
                sql.Identifier(schema),
                sql.Identifier(table)
            )
        else:
            from_clause = sql.Identifier(table_name)
        from_with_alias = sql.SQL('{} AS {}').format(from_clause, sql.Identifier(main_alias))

    joins = []
    joined_tables = dataset.tables.filter(joined_on_type__isnull=False).order_by('id')
    table_id_to_alias = {main_table.id: main_alias}

    for idx, join_table in enumerate(joined_tables, start=1):
        if not join_table.joined_on_left or not join_table.joined_on_right:
            continue
        table_alias = f't{idx}'
        table_id_to_alias[join_table.id] = table_alias
        is_join_file_source = join_table.file_upload_id is not None

        if is_join_file_source:
            try:
                df_join = read_file_to_dataframe(
                    join_table.file_upload_id,
                    sheet_name=getattr(join_table, 'sheet_name', None),
                    row_limit=None
                )
                join_table_columns[join_table.id] = list(df_join.columns)
                join_table_ref, _ = dataframe_to_sql_values(df_join, table_alias, row_limit=None)
            except Exception as e:
                raise ValueError(f"Ошибка чтения файла для JOIN: {str(e)}")
        else:
            table_name = join_table.table_name
            if '.' in table_name:
                schema, table = table_name.split('.', 1)
                join_from_clause = sql.SQL('{}.{}').format(
                    sql.Identifier(schema),
                    sql.Identifier(table)
                )
            else:
                join_from_clause = sql.Identifier(table_name)
            join_table_ref = sql.SQL('{} AS {}').format(join_from_clause, sql.Identifier(table_alias))

        join_type = (join_table.joined_on_type or 'LEFT JOIN').strip().upper()
        if 'JOIN' not in join_type:
            join_type = f'{join_type} JOIN'
        left_table_alias = main_alias

        left_col_ref = (
            f'col_{main_table_columns.index(join_table.joined_on_left)}'
            if is_file_source and main_table_columns and join_table.joined_on_left in main_table_columns
            else join_table.joined_on_left
        )
        right_col_ref = (
            f'col_{join_table_columns[join_table.id].index(join_table.joined_on_right)}'
            if is_join_file_source and join_table.joined_on_right in join_table_columns.get(join_table.id, [])
            else join_table.joined_on_right
        )
        join_condition = sql.SQL('{}.{} = {}.{}').format(
            sql.Identifier(left_table_alias),
            sql.Identifier(left_col_ref),
            sql.Identifier(table_alias),
            sql.Identifier(right_col_ref)
        )
        joins.append(sql.SQL('{} {} ON {}').format(
            sql.SQL(join_type),
            join_table_ref,
            join_condition
        ))
    
    # Собираем условия WHERE (та же логика, что и в build_dataset_query)
    where_conditions = []
    
    if where_clause:
        where_conditions.append(sql.SQL(where_clause))
    
    if search:
        def _count_col_ref(table_id, source_column):
            if table_id == main_table.id and main_table_columns and source_column in main_table_columns:
                return f'col_{main_table_columns.index(source_column)}'
            if table_id in join_table_columns and source_column in join_table_columns[table_id]:
                return f'col_{join_table_columns[table_id].index(source_column)}'
            return source_column

        search_conditions = []
        for field in dataset.fields.all():
            table_alias = table_id_to_alias.get(field.source_table.id, main_alias)
            col_ref = _count_col_ref(field.source_table.id, field.source_column)
            search_conditions.append(
                sql.SQL('CAST({}.{} AS TEXT) ILIKE {}').format(
                    sql.Identifier(table_alias),
                    sql.Identifier(col_ref),
                    sql.Literal(f'%{search}%')
                )
            )
        
        if search_conditions:
            where_conditions.append(
                sql.SQL('({})').format(sql.SQL(' OR ').join(search_conditions))
            )
    
    # Собираем итоговый COUNT запрос
    query_parts = [
        sql.SQL('SELECT COUNT(*)'),
        sql.SQL('FROM {}').format(from_with_alias)
    ]
    
    # Добавляем JOIN'ы
    query_parts.extend(joins)
    
    # Добавляем WHERE если есть условия
    if where_conditions:
        query_parts.append(sql.SQL('WHERE {}').format(sql.SQL(' AND ').join(where_conditions)))
    
    final_query = sql.SQL(' ').join(query_parts)
    return final_query


def get_dataset_table_alias(dataset, table_id):
    """Возвращает алиас таблицы в запросе датасета."""
    main_table = dataset.tables.filter(joined_on_type__isnull=True).first()
    if main_table and main_table.id == table_id:
        return 't0'
    
    joined_tables = dataset.tables.filter(joined_on_type__isnull=False).order_by('id')
    for idx, join_table in enumerate(joined_tables, start=1):
        if join_table.id == table_id:
            return f't{idx}'
    
    return 't0'  # По умолчанию



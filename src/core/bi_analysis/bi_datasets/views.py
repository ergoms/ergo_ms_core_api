from uuid import uuid4
from django.core.files import File
from django.db import transaction, connection, ProgrammingError
from django.shortcuts import get_object_or_404
from django.http import Http404
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from psycopg2 import sql
from .models import Connection

from rest_framework import (
    generics,
    permissions,
    status,
    viewsets
)
from rest_framework.exceptions import ValidationError
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView

from openpyxl import load_workbook
import tempfile, os, openpyxl, csv
import logging

logger = logging.getLogger(__name__)

from src.core.bi_analysis.bi_datasets.models import Dataset, FileUpload, DataSetTable, DataSetField, DatasetParam
from src.core.bi_analysis.bi_datasets.serializers import (
    DatasetDetailSerializer,
    DatasetSerializer,
    FileUploadSerializer,
    DataSetTableSerializer,
    DataSetFieldSerializer,
    DatasetShortSerializer, 
    DatasetUpdateSerializer,
    DatasetDetailFullSerializer,
    DatasetParamSerializer
)

from src.core.bi_analysis.services.services import (
    create_temp_table_from_source,
    import_file_upload_to_table,
    populate_initial_fields,
    populate_initial_fields_from_file,
    auto_join_table,
    introspect_columns,
    rebuild_dataset_joins,
    table_exists,
    build_dataset_query,
)

from ..bi_charts.methods import get_rows_for_chart


# ==============================================================================
# Dataset endpoints
# ==============================================================================

class DatasetRowsAPIView(APIView):
    def get(self, request, pk):
        dataset = get_object_or_404(Dataset, pk=pk, owner=request.user)
        chart_fields = dataset.fields.all()
        rows = get_rows_for_chart(dataset, chart_fields)
        return Response(rows)

class DatasetListCreateView(generics.ListCreateAPIView):
    queryset = Dataset.objects.all()
    serializer_class = DatasetSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            # Для генерации схемы Swagger возвращаем пустой queryset
            return Dataset.objects.none()
        return (Dataset.objects
            .filter(owner=self.request.user)
            .only('id', 'name', 'created_at', 'owner')
            .order_by('-created_at'))

    def get_serializer_class(self):
        # Для списка отдаём короткий сериализатор, чтобы не тянуть связанные таблицы/поля
        if self.request and self.request.method == 'GET':
            return DatasetShortSerializer
        return DatasetSerializer

    @transaction.atomic
    def perform_create(self, serializer):
        dataset = serializer.save(owner=self.request.user)
        if not dataset.file_source:
            raise ValidationError("file_source is required for dataset creation")

        # НЕ создаем таблицы в БД - используем метаданные и читаем файлы напрямую
        # table_ref оставляем пустым или используем имя файла для обратной совместимости
        dataset.table_ref = None  # Больше не нужна материализованная таблица
        dataset.save(update_fields=['table_ref'])

        # Создаем DataSetTable с ссылкой на file_upload
        # table_name может быть пустым или содержать имя для отображения
        main_table = DataSetTable.objects.create(
            dataset=dataset,
            connection=dataset.connection,
            table_name=dataset.file_source.original_filename or f"file_{dataset.file_source.id}",  # Имя для отображения
            joined_on={},
            file_upload=dataset.file_source 
        )
        
        if dataset.file_source and dataset.file_source.columns_info:
            main_table.columns_info = dataset.file_source.columns_info
            main_table.save(update_fields=["display_name", "columns_info"])

        # Создаем поля на основе информации о колонках из файла
        populate_initial_fields_from_file(dataset, dataset.file_source, main_table)
        
        # Безопасный доступ к request.data
        request_data = getattr(self.request, 'data', {})
        fields_data = request_data.get('fields', []) if isinstance(request_data, dict) else []
        if fields_data:
            for field in fields_data:
                obj = dataset.fields.filter(name=field.get('name')).first()
                if obj:
                    update_fields = []
                    if 'aggregation' in field:
                        obj.aggregation = field['aggregation']
                        update_fields.append('aggregation')
                    if 'type' in field:
                        obj.type = field['type']
                        update_fields.append('type')
                    if update_fields:
                        obj.save(update_fields=update_fields)
        params_data = request_data.get('params') if isinstance(request_data, dict) else None
        if params_data is not None:
            # поддержка формата из фронта: [{name,type,defaultValue,sourceUsage,...}]
            try:
                items = []
                for i, p in enumerate(params_data or []):
                    if not isinstance(p, dict):
                        continue
                    items.append({
                        'name': p.get('name'),
                        'type': p.get('type') or 'string',
                        'default_value': p.get('defaultValue', p.get('default')),
                        'source_usage': p.get('sourceUsage', False),
                        'order': p.get('order', i),
                        'description': p.get('description', ''),
                    })
                dataset.set_params_items(items)
            except Exception as e:
                # не прерываем создание датасета из-за параметров
                print(f"[DatasetListCreateView] params save failed: {e}")
        
class DatasetRemoveRelationView(APIView):
    """
    POST /bi_analysis/bi_datasets/<pk>/remove-relation/
    { "right_table_id": 1339 }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        request_data = getattr(request, 'data', {})
        right_id = request_data.get("right_table_id") if isinstance(request_data, dict) else None
        if not right_id:
            return Response(
                {"success": False, "error": "right_table_id is required"}, status=400
            )

        dataset = get_object_or_404(Dataset, pk=pk, owner=request.user)
        try:
            tbl = dataset.tables.get(pk=right_id)
        except DataSetTable.DoesNotExist:
            return Response(
                {"success": False, "error": "table not found in dataset"}, status=404
            )

        # Удаляем все поля из DataSetField, которые ссылаются на эту таблицу
        deleted_fields, _ = dataset.fields.filter(source_table=tbl).delete()

        # Очищаем связь
        tbl.joined_on_type = tbl.joined_on_left = tbl.joined_on_right = None
        tbl.joined_on = {}
        tbl.save()

        from ..services.services import rebuild_dataset_joins
        rebuild_dataset_joins(dataset)

        serializer = DatasetDetailSerializer(dataset)
        return Response({"success": True, "dataset": serializer.data})


class DatasetListView(generics.ListAPIView):
    serializer_class = DatasetShortSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            # Для генерации схемы Swagger возвращаем пустой queryset
            return Dataset.objects.none()
        return (Dataset.objects
            .filter(owner=self.request.user)
            .order_by('-created_at'))

class DatasetDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Dataset.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            # Для генерации схемы Swagger возвращаем пустой queryset
            return Dataset.objects.none()
        return Dataset.objects.filter(owner=self.request.user)

    def get_serializer_class(self):
        if self.request.method in ('PUT', 'PATCH'):
            return DatasetUpdateSerializer
        return DatasetDetailFullSerializer
    
class DataSetTableColumnsView(APIView):
    def get(self, request, pk):
        # pk — это id DataSetTable
        try:
            tbl = DataSetTable.objects.get(pk=pk)
        except DataSetTable.DoesNotExist:
            return Response({'error': 'Table not found'}, status=404)
        # Получаем все DataSetField, относящиеся к этой таблице
        fields = DataSetField.objects.filter(source_table=tbl)
        columns = [f.source_column or f.name for f in fields]
        return Response({'columns': columns})
    

class DatasetJoinTableView(APIView):
    """
    Прицепить (или переприцепить) staging-таблицу к датасету.
    POST body:
        {
            "staging_name": "temp_abcd1234…",
            "left_column" : "Город",
            "right_column": "Город",
            "join_type"   : "INNER JOIN"   # опционально, default — INNER JOIN
        }
    """
    permission_classes = []

    def post(self, request, pk, *args, **kwargs):
        dataset = get_object_or_404(Dataset, pk=pk)

        request_data = getattr(request, 'data', {})
        if not isinstance(request_data, dict):
            request_data = {}
        staging_name = request_data.get("staging_name") or request_data.get("stagingName")
        left_column  = request_data.get("left_column")  or request_data.get("leftColumn")
        right_column = request_data.get("right_column") or request_data.get("rightColumn")
        join_type    = (request_data.get("join_type")   or
                        request_data.get("joinType")    or
                        "INNER JOIN").upper()

        if not all([staging_name, left_column, right_column]):
            return Response(
                {"success": False, "error": "staging_name, left_column и right_column обязательны"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            auto_join_table(
                dataset        = dataset,
                staging_name   = staging_name,
                left_column    = left_column,
                right_column   = right_column,
                join_type      = join_type,
            )
        except Exception as exc:
            return Response(
                {"success": False, "error": str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response({"success": True})

class DatasetAddRelationView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, dataset_id):
        from src.core.bi_analysis.bi_datasets.models import DataSetTable, FileUpload

        dataset = Dataset.objects.get(id=dataset_id)
        request_data = getattr(request, 'data', {})
        if not isinstance(request_data, dict):
            request_data = {}
        right_table_id = int(request_data.get('rightTableId', 0))
        join_type = request_data.get('joinType')
        lines = request_data.get('lines', [])
        if not lines:
            return Response({'error': 'Нет пар колонок для соединения'}, status=400)
        left_col = lines[0]['left']
        right_col = lines[0]['right']

        file_upload = None

        request_data = getattr(request, 'data', {})
        file_id = request_data.get("file_id") if isinstance(request_data, dict) else None
        if file_id:
            try:
                file_upload = FileUpload.objects.get(pk=file_id, owner=request.user)
            except FileUpload.DoesNotExist:
                return Response({'error': f'Файл с id={file_id} не найден или не принадлежит вам.'}, status=404)
        else:
            try:
                ds_table = DataSetTable.objects.get(pk=right_table_id, dataset__owner=request.user)
                file_upload = ds_table.file_upload
            except DataSetTable.DoesNotExist:
                try:
                    file_upload = FileUpload.objects.get(pk=abs(right_table_id), owner=request.user)
                except FileUpload.DoesNotExist:
                    return Response({'error': f'Не найден ни DataSetTable, ни FileUpload с id={right_table_id}'}, status=404)
        existing = dataset.tables.filter(pk=right_table_id).first()
        if existing:
            existing.joined_on_type  = join_type.upper()
            existing.joined_on_left  = left_col
            existing.joined_on_right = right_col
            existing.save(update_fields=["joined_on_type", "joined_on_left", "joined_on_right"])
            return Response({"success": True})

        if not file_upload:
            return Response({'error': f'Файл для связи с id={right_table_id} не найден.'}, status=404)

        tbl = DataSetTable.objects.create(
            dataset=dataset,
            connection=dataset.connection,
            file_upload=file_upload,
            display_name=file_upload.original_filename,
            columns_info=file_upload.columns_info,
            table_name=f"temp_{uuid4().hex}",
            joined_on_type=join_type.upper(),
            joined_on_left=left_col,
            joined_on_right=right_col,
        )
        
        rebuild_dataset_joins(dataset)
        return Response({'success': True})
        
class AddTableToDatasetView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        request_data = getattr(request, 'data', {})
        file_id = request_data.get('file_id') if isinstance(request_data, dict) else None
        dataset = get_object_or_404(Dataset, pk=pk, owner=request.user)

        try:
            file_upload = FileUpload.objects.get(pk=file_id, owner=request.user)
        except FileUpload.DoesNotExist:
            return Response({'error': f'Файл с id={file_id} не найден или не принадлежит вам.'}, status=404)
        
        connection = file_upload.connection or dataset.connection

        # НЕ создаем таблицы в БД - используем только метаданные
        # table_name используется только для отображения/идентификации
        table_display_name = file_upload.original_filename or f"file_{file_upload.id}"

        data_table = DataSetTable.objects.create(
            dataset=dataset,
            connection=connection,
            table_name=table_display_name,  # Имя для отображения, не реальная таблица
            joined_on={},
            file_upload=file_upload
        )

        return Response({
            "id": data_table.id,
            "table_ref": None,  # Больше не используется
            "name": data_table.table_name,
            "display_name": file_upload.original_filename,
        })

# ==============================================================================
# DatasetViewSet (альтернатива generic views)
# ==============================================================================
    
class DatasetPreviewView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        dataset = Dataset.objects.filter(pk=pk, owner=request.user).first()
        if not dataset:
            return Response({"detail": "Not found"}, status=404)

        limit = int(request.query_params.get('limit', 1000000))
        offset = int(request.query_params.get('offset', 0))
        search = request.query_params.get('search', '').strip()
        use_async = request.query_params.get('async', 'false').lower() == 'true'
        
        try:
            # Проверяем, есть ли таблицы в датасете
            if not dataset.tables.exists():
                return Response({"detail": "Таблицы не найдены в датасете. Добавьте главную таблицу для предпросмотра."}, status=400)
            
            # Получаем главную таблицу
            main_table = dataset.tables.filter(joined_on_type__isnull=True).first()
            if not main_table:
                return Response({"detail": "Не найдена главная таблица для датасета"}, status=400)
            
            # Проверяем, есть ли JOIN'ы в датасете
            has_joins = dataset.tables.filter(joined_on_type__isnull=False).exists()
            
            # Для простого случая (одна таблица, файловый источник, без JOIN'ов и поиска) используем прямое чтение файла
            # Это намного быстрее, чем конвертация в SQL VALUES clause или асинхронная обработка
            is_file_source = main_table.file_upload_id is not None
            use_fast_path = is_file_source and not has_joins and not search and offset == 0
            
            # Для быстрого пути НЕ используем асинхронную обработку - это намного быстрее напрямую
            # Асинхронная обработка нужна только для сложных случаев (JOIN'ы, поиск, очень большие файлы)
            if use_fast_path:
                # Быстрый путь: прямое чтение файла через polars, как в upload эндпоинте
                # ВАЖНО: учитываем поля датасета (dataset.fields) для выбора и порядка колонок,
                # чтобы предпросмотр соответствовал настроенным полям, а не всем колонкам файла.
                try:
                    from src.core.bi_analysis.services.services import read_file_to_dataframe, count_file_rows
                    
                    # Подсчитываем общее количество строк (только при первой загрузке или если offset=0)
                    total_count = None
                    if offset == 0:
                        try:
                            total_count = count_file_rows(
                                main_table.file_upload_id,
                                sheet_name=getattr(main_table, 'sheet_name', None)
                            )
                        except Exception as e:
                            logger.warning(f"Не удалось подсчитать общее количество строк: {str(e)}")
                    
                    # Читаем файл с лимитом строк
                    table_data = read_file_to_dataframe(
                        main_table.file_upload_id,
                        sheet_name=getattr(main_table, 'sheet_name', None),
                        row_limit=limit
                    )

                    # Если в датасете определены поля, используем их для выбора и порядка колонок.
                    # Это делает быстрый путь семантически эквивалентным SQL-пути, который
                    # строится через build_dataset_query и использует dataset.fields.order_by('order').
                    dataset_fields_qs = dataset.fields.all().order_by('order')
                    if dataset_fields_qs.exists():
                        # Сначала пытаемся выбрать колонки исходя из source_column (если заданы),
                        # при отсутствии — по имени поля (name). Отсутствующие в файле колонки пропускаем.
                        selected_columns = []
                        for f in dataset_fields_qs:
                            src_col = f.source_column or f.name
                            if src_col and src_col in table_data.columns:
                                selected_columns.append(src_col)

                        if selected_columns:
                            # Фильтруем и переупорядочиваем колонки в соответствии с полями датасета
                            table_data = table_data.select(selected_columns)

                    # Конвертируем в формат ответа
                    columns = table_data.columns
                    rows = table_data.rows
                    
                    # Преобразуем строки в список кортежей (как в SQL результате)
                    rows_tuples = [tuple(row) for row in rows]
                    
                    response_data = {
                        "columns": columns,
                        "rows": rows_tuples,
                        "has_more": len(rows) == limit  # Указываем, есть ли еще данные
                    }
                    
                    # Добавляем total_count, если он был подсчитан
                    if total_count is not None:
                        response_data["total_count"] = total_count
                    
                    return Response(response_data)
                except Exception as e:
                    # Если не удалось использовать быстрый путь, используем обычный
                    logger.warning(f"Не удалось использовать быстрый путь для preview: {str(e)}. Используется SQL путь.")
            else:
                # Для сложных случаев (с JOIN'ами, поиском, offset) проверяем нужна ли асинхронная обработка
                # Асинхронная обработка только для действительно больших лимитов или при явном запросе
                # Увеличиваем порог, чтобы для обычных случаев (до 5000 строк) использовался синхронный режим
                async_threshold = 5000
                # Проверяем limit только если он указан (не None)
                limit_exceeds_threshold = limit is not None and limit > async_threshold
                should_use_async = use_async or (limit_exceeds_threshold and has_joins)
                
                if should_use_async:
                    try:
                        from src.core.bi_analysis.tasks import process_dataset_preview
                        task = process_dataset_preview.delay(pk, limit)
                        return Response({
                            "task_id": task.id,
                            "status": "processing",
                            "message": "Предпросмотр обрабатывается асинхронно"
                        }, status=202)
                    except Exception as e:
                        # Если Celery недоступен, используем синхронный режим
                        logger.warning(f"Не удалось запустить асинхронную задачу Celery: {str(e)}. Используется синхронный режим.")
                        # Продолжаем выполнение в синхронном режиме ниже
            
            # Обычный путь: через SQL запрос (для JOIN'ов, поиска, offset и т.д.)
            from src.core.bi_analysis.services.services import build_dataset_query, build_dataset_count_query
            
            # Подсчитываем общее количество строк (только при первой загрузке или если offset=0)
            total_count = None
            if offset == 0:
                try:
                    count_query = build_dataset_count_query(
                        dataset,
                        search=search if search else None
                    )
                    with connection.cursor() as count_cursor:
                        count_cursor.execute(count_query)
                        total_count = count_cursor.fetchone()[0]
                except Exception as e:
                    logger.warning(f"Не удалось подсчитать общее количество строк: {str(e)}")
            
            # При поиске не применяем лимит, чтобы поиск выполнялся по всем данным
            # Без поиска используем лимит для пагинации
            query_limit = None if search else limit
            
            query, display_columns = build_dataset_query(
                dataset,
                limit=query_limit,
                offset=offset,
                search=search if search else None
            )

            with connection.cursor() as cursor:
                cursor.execute(query)
                rows = cursor.fetchall()
                columns = display_columns if display_columns else [col[0] for col in cursor.description]
                    
        except ValueError as e:
            return Response({"detail": str(e)}, status=400)
        except ProgrammingError as e:
            return Response({"detail": f"Ошибка выполнения запроса: {str(e)}"}, status=500)
        except Exception as e:
            return Response({"detail": f"Ошибка: {str(e)}"}, status=500)

        response_data = {
            "columns": columns,
            "rows": rows,
            "has_more": len(rows) == limit  # Указываем, есть ли еще данные
        }
        
        # Добавляем total_count, если он был подсчитан
        if total_count is not None:
            response_data["total_count"] = total_count
        
        return Response(response_data)


class DatasetPreviewTaskStatusView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        Получает статус и результат асинхронной задачи предпросмотра.
        """
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({"detail": "Не указан task_id"}, status=400)
        
        try:
            from celery.result import AsyncResult
            from src.config.celery import celery_app
            
            task = AsyncResult(task_id, app=celery_app)
            
            if task.state == 'PENDING':
                response = {
                    'task_id': task_id,
                    'status': 'pending',
                    'message': 'Задача ожидает выполнения'
                }
            elif task.state == 'PROGRESS':
                response = {
                    'task_id': task_id,
                    'status': 'processing',
                    'message': 'Задача выполняется',
                    'progress': task.info.get('progress', 0) if isinstance(task.info, dict) else None
                }
            elif task.state == 'SUCCESS':
                result = task.result
                response = {
                    'task_id': task_id,
                    'status': 'success',
                    'result': result
                }
            elif task.state == 'FAILURE':
                response = {
                    'task_id': task_id,
                    'status': 'failure',
                    'error': str(task.info) if task.info else 'Неизвестная ошибка'
                }
            else:
                response = {
                    'task_id': task_id,
                    'status': task.state.lower(),
                    'message': f'Статус задачи: {task.state}'
                }
            
            return Response(response, status=200)
            
        except Exception as e:
            return Response({"detail": f"Ошибка получения статуса задачи: {str(e)}"}, status=500)


class DatasetColumnsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        dataset = Dataset.objects.filter(pk=pk, owner=request.user).first()
        if not dataset:
            return Response({"detail": "Dataset not found"}, status=404)

        # Получаем поля датасета с их типами
        fields = dataset.fields.all()
        cols = []
        for f in fields:
            cols.append({
                "id": f.id,
                "name": f.name,
                "type": f.type,
                "aggregation": f.aggregation,
            })

        return Response({"columns": cols})

class DatasetDraftPreviewView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        Принимает черновик датасета (главная таблица + join'ы), возвращает предпросмотр данных.
        Работает без создания постоянных таблиц - читает файлы напрямую через polars.
        Поддерживает асинхронную обработку через celery для больших файлов.
        """
        from src.core.bi_analysis.services.services import read_file_to_dataframe, dataframe_to_sql_values
        from psycopg2 import sql
        
        data = request.data
        
        # Проверяем наличие главной таблицы
        main_table = data.get('mainTable')
        if not main_table or (isinstance(main_table, dict) and not main_table):
            return Response({
                "error": "Не указана главная таблица",
                "columns": [],
                "rows": []
            }, status=400)
        
        from django.conf import settings
        
        use_async = data.get('async', False)
        # Лимиты убраны - если limit не указан, загружаем все данные
        # Если limit указан, используем его для пагинации
        limit = data.get('limit')
        if limit is not None:
            limit = int(limit)
        
        # Для больших лимитов или множества файлов используем асинхронную обработку
        file_count = sum([
            1 if 'file_id' in data.get('mainTable', {}) and data.get('mainTable', {}).get('file_id') else 0,
            sum([1 if 'file_id' in jt and jt.get('file_id') else 0 for jt in data.get('joinedTables', [])])
        ])
        
        async_threshold = getattr(settings, 'BI_PREVIEW_ASYNC_THRESHOLD', 5000)
        # Проверяем limit только если он указан (не None)
        limit_exceeds_threshold = limit is not None and limit > async_threshold
        if use_async or limit_exceeds_threshold or file_count > 2:
            try:
                from src.core.bi_analysis.tasks import process_draft_preview
                task = process_draft_preview.delay(data)
                return Response({
                    "task_id": task.id,
                    "status": "processing",
                    "message": "Предпросмотр обрабатывается асинхронно"
                }, status=202)
            except Exception as e:
                # Если Celery недоступен, используем синхронную обработку
                logger.warning(f"Не удалось запустить асинхронную задачу Celery: {e}. Используем синхронную обработку.")
                # Продолжаем выполнение синхронно
        
        connection_id = data.get('connection_id')
        joined_tables = data.get('joinedTables', [])
        offset = int(data.get('offset', 0))

        def get_sheet_name(table_dict):
            if not isinstance(table_dict, dict):
                return None
            sheet = (
                table_dict.get('sheet_name') or
                table_dict.get('sheetName') or
                table_dict.get('sheet')
            )
            if sheet:
                return sheet
            nested_file = table_dict.get('file_upload') or {}
            if isinstance(nested_file, dict):
                return nested_file.get('sheet_name') or nested_file.get('sheet')
            return None

        # Определяем, является ли главная таблица файловым источником
        is_main_file = 'file_id' in main_table and main_table.get('file_id')
        
        # Формируем FROM для главной таблицы
        main_alias = 'a'
        
        if is_main_file:
            # Для файловых источников читаем данные напрямую
            # Вычисляем сколько строк нужно прочитать: limit + offset + запас для корректной работы
            # Ограничиваем максимальное количество строк в VALUES clause для производительности
            file_read_limit = None
            if limit is not None:
                # Читаем достаточно данных с учетом offset и небольшого запаса
                # Убрано ограничение MAX_VALUES_ROWS - загружаем все строки
                file_read_limit = limit + (offset or 0) + 100  # Запас в 100 строк для корректной работы
            
            try:
                df_main = read_file_to_dataframe(
                    main_table['file_id'],
                    sheet_name=get_sheet_name(main_table),
                    row_limit=file_read_limit
                )
                
                # Проверяем, что данные прочитаны
                if not df_main or not hasattr(df_main, 'columns') or not df_main.columns or len(df_main) == 0:
                    logger.warning(f"Файл {main_table['file_id']} прочитан, но не содержит данных")
                    return Response({
                        "columns": [],
                        "rows": [],
                        "error": "Файл не содержит данных или выбран неправильный лист"
                    }, status=200)
                
                # Передаем None для row_limit в dataframe_to_sql_values, чтобы включить все прочитанные данные
                # Лимит и offset будут применены в SQL запросе через LIMIT/OFFSET clauses
                main_from, _ = dataframe_to_sql_values(df_main, main_alias, row_limit=None)
                main_cols = [str(col) for col in df_main.columns]
                
                logger.info(f"Прочитано {len(df_main)} строк из файла {main_table['file_id']}, колонок: {len(main_cols)}")
            except Exception as e:
                logger.error(f"Ошибка чтения главного файла {main_table.get('file_id')}: {str(e)}", exc_info=True)
                raise ValidationError(f"Ошибка чтения главного файла: {str(e)}")
        else:
            # Для БД источников используем прямое подключение
            table_name = main_table.get('table_name')
            if not table_name:
                raise ValidationError("Не указано имя таблицы для БД источника")
            
            if '.' in table_name:
                schema, table = table_name.split('.', 1)
                main_from = sql.SQL('{}.{} AS {}').format(
                    sql.Identifier(schema),
                    sql.Identifier(table),
                    sql.Identifier(main_alias)
                )
            else:
                main_from = sql.SQL('{} AS {}').format(
                    sql.Identifier(table_name),
                    sql.Identifier(main_alias)
                )
            
            # Получаем колонки из БД таблицы
            main_cols = introspect_columns(table_name)
        
        # Формируем SELECT с уникальными алиасами (col_0, col_1, ...), чтобы избежать
        # неоднозначности при длинных именах колонок: в PostgreSQL идентификаторы обрезаются
        # до 63 байт (кириллица = 2 байта → ~31 символ), из‑за чего разные колонки могут
        # получить один и тот же алиас. Для файлового источника подзапрос отдаёт col_0, col_1, ...
        display_column_names = []
        select_parts = []
        col_idx = 0
        for col in main_cols:
            col_ref = f'col_{col_idx}' if is_main_file else col
            select_parts.append(sql.SQL('{}.{} AS {}').format(
                sql.Identifier(main_alias),
                sql.Identifier(col_ref),
                sql.Identifier(f'col_{col_idx}')
            ))
            display_column_names.append(col)
            col_idx += 1

        # Обрабатываем JOIN'ы
        join_clauses = []
        alias_idx = ord('b')
        all_cols = set(main_cols)
        
        for jt in joined_tables:
            tbl_alias = chr(alias_idx)
            alias_idx += 1
            
            # Определяем, является ли JOIN таблица файловым источником
            is_join_file = 'file_id' in jt and jt.get('file_id')
            
            if is_join_file:
                # Для файловых источников читаем данные напрямую
                # Для JOIN таблиц читаем все данные, лимит будет применен к итоговому запросу
                try:
                    df_join = read_file_to_dataframe(
                        jt['file_id'],
                        sheet_name=get_sheet_name(jt),
                        row_limit=None  # Читаем все данные для JOIN
                    )
                    join_from, _ = dataframe_to_sql_values(df_join, tbl_alias, row_limit=None)
                    join_cols = [str(col) for col in df_join.columns]
                except Exception as e:
                    raise ValidationError(f"Ошибка чтения файла для JOIN: {str(e)}")
            else:
                # Для БД источников используем прямое подключение
                table_name = jt.get('table_name')
                if not table_name:
                    raise ValidationError("Не указано имя таблицы для JOIN БД источника")
                
                if '.' in table_name:
                    schema, table = table_name.split('.', 1)
                    join_from = sql.SQL('{}.{} AS {}').format(
                        sql.Identifier(schema),
                        sql.Identifier(table),
                        sql.Identifier(tbl_alias)
                    )
                else:
                    join_from = sql.SQL('{} AS {}').format(
                        sql.Identifier(table_name),
                        sql.Identifier(tbl_alias)
                    )
                
                join_cols = introspect_columns(table_name)
            
            # Добавляем колонки из JOIN таблицы (только уникальные) с уникальными алиасами
            for j, col in enumerate(join_cols):
                if col not in all_cols:
                    join_col_ref = f'col_{j}' if is_join_file else col
                    select_parts.append(sql.SQL('{}.{} AS {}').format(
                        sql.Identifier(tbl_alias),
                        sql.Identifier(join_col_ref),
                        sql.Identifier(f'col_{col_idx}')
                    ))
                    display_column_names.append(col)
                    col_idx += 1
                    all_cols.add(col)
            
            # Формируем условие JOIN
            join_type = jt.get('joinType', 'LEFT JOIN').strip().upper()
            if 'JOIN' not in join_type:
                join_type = f'{join_type} JOIN'
            
            lines = jt.get('lines') or []
            if not lines or not lines[0].get('left') or not lines[0].get('right'):
                raise ValidationError("Не указаны поля для join'а")
            
            left_col = lines[0]['left']
            right_col = lines[0]['right']
            left_col_ref = f'col_{main_cols.index(left_col)}' if is_main_file and left_col in main_cols else left_col
            right_col_ref = f'col_{join_cols.index(right_col)}' if is_join_file and right_col in join_cols else right_col
            
            join_condition = sql.SQL('{}.{} = {}.{}').format(
                sql.Identifier(main_alias),
                sql.Identifier(left_col_ref),
                sql.Identifier(tbl_alias),
                sql.Identifier(right_col_ref)
            )
            
            join_clause = sql.SQL('{} {} ON {}').format(
                sql.SQL(join_type),
                join_from,
                join_condition
            )
            join_clauses.append(join_clause)
        
        # Проверяем, что есть колонки для SELECT
        if not select_parts:
            logger.warning("Нет колонок для SELECT в запросе")
            return Response({
                "columns": [],
                "rows": [],
                "error": "Нет колонок для отображения"
            }, status=200)
        
        # Собираем итоговый запрос
        query_parts = [
            sql.SQL('SELECT {}').format(sql.SQL(', ').join(select_parts)),
            sql.SQL('FROM {}').format(main_from)
        ]
        
        query_parts.extend(join_clauses)
        
        # Добавляем OFFSET если есть
        if offset is not None and offset > 0:
            query_parts.append(sql.SQL('OFFSET {}').format(sql.Literal(offset)))
        
        # Добавляем LIMIT
        if limit is not None:
            query_parts.append(sql.SQL('LIMIT {}').format(sql.Literal(limit)))
        
        final_query = sql.SQL(' ').join(query_parts)
        
        # Выполняем запрос
        try:
            logger.info(f"Выполняем SQL запрос с limit={limit}, offset={offset}")
            with connection.cursor() as cursor:
                cursor.execute(final_query)
                rows = cursor.fetchall()
            
            logger.info(f"Получено {len(rows)} строк, {len(display_column_names)} колонок")

            return Response({
                "columns": display_column_names,
                "rows": rows,
                "has_more": len(rows) == limit if limit else False
            })
        except Exception as e:
            logger.error(f"Ошибка выполнения SQL запроса в draftPreview: {str(e)}", exc_info=True)
            # Возвращаем более информативную ошибку
            return Response({
                "error": f"Ошибка выполнения запроса: {str(e)}",
                "detail": "Проверьте, что файл содержит данные и правильно выбран лист для Excel файлов",
                "columns": [],
                "rows": []
            }, status=500)

# ==============================================================================
# DataSetTable endpoints
# ==============================================================================

class DataSetTableViewSet(viewsets.ModelViewSet):
    queryset = DataSetTable.objects.all()
    serializer_class = DataSetTableSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save()
    
class RenameDatasetColumnsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        dataset = Dataset.objects.get(pk=pk, owner=request.user)
        request_data = getattr(request, 'data', {})
        renames = request_data.get('renames', []) if isinstance(request_data, dict) else []
        table_ref = dataset.table_ref
        schema, table = ('public', table_ref)
        if '.' in table_ref:
            schema, table = table_ref.split('.', 1)

        with connection.cursor() as cursor:
            for rename in renames:
                cursor.execute(
                    f'ALTER TABLE "{schema}"."{table}" RENAME COLUMN "{rename["old_name"]}" TO "{rename["new_name"]}";'
                )
                DataSetField.objects.filter(
                    dataset=dataset, source_column=rename["old_name"]
                ).update(name=rename["new_name"], source_column=rename["new_name"])
        return Response({"status": "ok"})

# ==============================================================================
# DataSetField endpoints
# ==============================================================================

class DataSetFieldViewSet(viewsets.ModelViewSet):
    serializer_class = DataSetFieldSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        dataset_id = self.request.query_params.get('dataset')
        queryset = DataSetField.objects.all()
        if dataset_id:
            queryset = queryset.filter(dataset_id=dataset_id)
        return queryset

# ==============================================================================
# DatasetParam endpoints
# ==============================================================================

from src.core.utils.mixins import SwaggerSafeMixin

class DatasetParamViewSet(SwaggerSafeMixin, viewsets.ModelViewSet):
    serializer_class = DatasetParamSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        base = DatasetParam.objects.select_related('dataset')
        user = self.get_safe_user()
        if user is None:
            return base.none()
        qs = base.filter(dataset__owner=user)
        dataset_id = self.request.query_params.get('dataset')
        if dataset_id:
            qs = qs.filter(dataset_id=dataset_id)
        return qs.order_by('order', 'id')

# ==============================================================================
# FileUpload endpoints
# ==============================================================================

class TempUploadView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Файл не передан'}, status=400)
        suffix = os.path.splitext(file.name)[-1]
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            for chunk in file.chunks():
                tmp.write(chunk)
            temp_path = tmp.name
        return Response({
            'temp_path': temp_path,
            'original_filename': file.name,
            'file_type': suffix.lstrip('.').lower()
        }, status=201)

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            # Для генерации схемы Swagger возвращаем пустой queryset
            return FileUpload.objects.none()
        return FileUpload.objects.filter(owner=self.request.user)

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user)

class FileUploadDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    GET    /upload/{pk}/    — метаданные + предпросмотр содержимого
    PUT    /upload/{pk}/    — изменить имя/загрузить новый файл
    DELETE /upload/{pk}/    — удалить запись и файл на диске
    """
    serializer_class = FileUploadSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            # Для генерации схемы Swagger возвращаем пустой queryset
            return FileUpload.objects.none()
        return FileUpload.objects.filter(owner=self.request.user)
    
    def perform_create(self, serializer):
        instance = serializer.save(owner=self.request.user)
        columns_info = extract_columns_info(instance)
        instance.columns_info = columns_info
        instance.save(update_fields=['columns_info'])

    def retrieve(self, request, *args, **kwargs):
        # get_object() уже обрабатывает DoesNotExist и возвращает Http404
        # Перехватываем только Http404 и DoesNotExist
        pk = kwargs.get('pk')
        logger.info(f"[RETRIEVE] Попытка получить FileUpload с ID={pk}, пользователь={request.user.id}")
        
        try:
            instance = self.get_object()
            logger.info(f"[RETRIEVE] FileUpload найден: ID={instance.id}, UUID={instance.file_uuid}, owner={instance.owner.id}")
        except FileUpload.DoesNotExist:
            logger.warning(f"[RETRIEVE] FileUpload с ID={pk} не найден (DoesNotExist), пользователь={request.user.id}")
            return Response(
                {"detail": "No FileUpload matches the given query."}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Http404:
            logger.warning(f"[RETRIEVE] FileUpload с ID={pk} не найден (Http404), пользователь={request.user.id}")
            return Response(
                {"detail": "No FileUpload matches the given query."}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        serializer = self.get_serializer(instance)
        data = serializer.data

        if not instance.file or not hasattr(instance.file, 'path'):
            data['parsed'] = []
            return Response(data)

        encoding   = request.query_params.get('encoding', 'utf-8')
        delimiter  = request.query_params.get('delimiter', ',')
        has_header = request.query_params.get('has_header', 'true').lower() == 'true'
        
        # Пагинация строк
        limit = request.query_params.get('limit')
        offset = request.query_params.get('offset', '0')
        limit = int(limit) if limit else None
        offset = int(offset) if offset else 0
        
        # Подсчитываем общее количество строк для пагинации
        total_rows_count = None
        try:
            from src.core.bi_analysis.services.services import count_file_rows
            sheet_name = request.query_params.get('sheet_name') or request.query_params.get('sheet')
            total_rows_count = count_file_rows(instance.id, sheet_name)
        except Exception as e:
            logger.warning(f"Не удалось подсчитать строки для файла {instance.id}: {str(e)}")

        try:
            # Проверяем, является ли файл бинарным
            from src.core.bi_analysis.bi_datasets.binary_storage import is_binary_file, read_from_binary
            
            if is_binary_file(instance.file.path) or instance.file_type == 'bin':
                # Читаем из бинарного файла
                columns, rows = read_from_binary(instance.file.path, row_limit=None)
                
                # Применяем пагинацию
                total_count = len(rows)
                if limit is not None:
                    paginated_rows = rows[offset:offset + limit]
                else:
                    paginated_rows = rows[offset:]
                
                if has_header and paginated_rows:
                    parsed = [columns, *paginated_rows]
                else:
                    parsed = paginated_rows
                
                data['parsed'] = parsed
                if total_rows_count is None:
                    data['total_count'] = total_count
                else:
                    data['total_count'] = total_rows_count
                return Response(data)
            
            if instance.file_type in ('csv', 'txt'):
                # Пробуем прочитать через бинарный формат, если файл конвертирован
                if is_binary_file(instance.file.path) or instance.file_type == 'bin':
                    columns, rows = read_from_binary(instance.file.path, row_limit=None)
                    if has_header and rows:
                        parsed = [columns, *rows]
                    else:
                        parsed = rows
                else:
                    # Читаем через Polars для эффективной обработки больших CSV
                    try:
                        import polars as pl
                        import math
                        
                        # Пробуем разные кодировки
                        df = None
                        try:
                            df = pl.read_csv(instance.file.path, encoding='utf8', separator=delimiter, try_parse_dates=True)
                        except Exception:
                            try:
                                df = pl.read_csv(instance.file.path, encoding='cp1251', separator=delimiter, try_parse_dates=True)
                            except Exception as e:
                                logger.error(f"Ошибка при чтении CSV через Polars: {e}")
                                # Fallback на старый метод
                                parsed = self._parse_csv(instance.file.path, encoding, delimiter, has_header)
                                df = None
                        
                        if df is not None:
                            # Конвертируем в список строк, заменяя NaN на None
                            # Используем to_numpy().tolist() для быстрой векторизованной конвертации (как для Excel)
                            columns = df.columns
                            rows_numpy = df.to_numpy().tolist()
                            
                            # Очищаем NaN и infinity значения
                            rows_list = []
                            for row in rows_numpy:
                                cleaned_row = []
                                for value in row:
                                    if value is None:
                                        cleaned_row.append(None)
                                    elif isinstance(value, float):
                                        if math.isnan(value) or math.isinf(value):
                                            cleaned_row.append(None)
                                        else:
                                            cleaned_row.append(value)
                                    else:
                                        cleaned_row.append(value)
                                rows_list.append(cleaned_row)
                            
                            # Применяем пагинацию
                            total_count = len(rows_list)
                            if limit is not None:
                                paginated_rows = rows_list[offset:offset + limit]
                            else:
                                paginated_rows = rows_list[offset:]
                            
                            if has_header:
                                parsed = [list(columns), *paginated_rows]
                            else:
                                parsed = paginated_rows
                            
                            # Добавляем общее количество строк
                            if total_rows_count is None:
                                data['total_count'] = total_count
                            else:
                                data['total_count'] = total_rows_count
                    except ImportError:
                        # Polars не установлен, используем старый метод
                        all_parsed = self._parse_csv(instance.file.path, encoding, delimiter, has_header)
                        
                        # Применяем пагинацию
                        if has_header and len(all_parsed) > 1:
                            header = [all_parsed[0]]
                            rows = all_parsed[1:]
                            total_count = len(rows)
                            if limit is not None:
                                paginated_rows = rows[offset:offset + limit]
                            else:
                                paginated_rows = rows[offset:]
                            parsed = header + paginated_rows
                            if total_rows_count is None:
                                data['total_count'] = total_count
                            else:
                                data['total_count'] = total_rows_count
                        else:
                            total_count = len(all_parsed)
                            if limit is not None:
                                parsed = all_parsed[offset:offset + limit]
                            else:
                                parsed = all_parsed[offset:]
                            if total_rows_count is None:
                                data['total_count'] = total_count
                            else:
                                data['total_count'] = total_rows_count

            elif instance.file_type == 'xlsx':
                # Читаем через Polars для эффективной обработки больших Excel файлов
                # Используем ту же логику, что и для подключений
                try:
                    import polars as pl
                    import math
                    
                    # Определяем лист для чтения
                    sheet_name = request.query_params.get('sheet_name') or request.query_params.get('sheet')
                    
                    # Читаем Excel файл через polars
                    if sheet_name:
                        df = pl.read_excel(instance.file.path, sheet_name=sheet_name)
                    else:
                        # Берем первый лист
                        df = pl.read_excel(instance.file.path, sheet_index=0)
                    
                    # Получаем список всех листов для возврата
                    # Используем openpyxl для получения списка листов (polars не предоставляет простой способ)
                    try:
                        wb = load_workbook(filename=instance.file.path, read_only=True)
                        sheets = wb.sheetnames
                        wb.close()
                    except:
                        sheets = []
                    
                    data['sheets'] = sheets
                    
                    # Конвертируем в список строк, заменяя NaN на None
                    # Используем to_numpy().tolist() для быстрой конвертации (векторизация polars)
                    columns = df.columns
                    rows_numpy = df.to_numpy().tolist()
                    
                    # Очищаем NaN и infinity значения
                    rows_list = []
                    for row in rows_numpy:
                        cleaned_row = []
                        for value in row:
                            if value is None:
                                cleaned_row.append(None)
                            elif isinstance(value, float):
                                if math.isnan(value) or math.isinf(value):
                                    cleaned_row.append(None)
                                else:
                                    cleaned_row.append(value)
                            else:
                                cleaned_row.append(value)
                        rows_list.append(cleaned_row)
                    
                    # Применяем пагинацию
                    total_count = len(rows_list)
                    if limit is not None:
                        paginated_rows = rows_list[offset:offset + limit]
                    else:
                        paginated_rows = rows_list[offset:]
                    
                    if has_header and paginated_rows:
                        parsed = [list(columns), *paginated_rows]
                    else:
                        parsed = paginated_rows
                    
                    # Добавляем общее количество строк
                    if total_rows_count is None:
                        data['total_count'] = total_count
                    else:
                        data['total_count'] = total_rows_count
                        
                except ImportError:
                    # Polars не установлен, используем старый метод
                    all_parsed, sheets = self._parse_xlsx(instance.file.path, has_header)
                    data['sheets'] = sheets
                    
                    # Применяем пагинацию
                    if has_header and len(all_parsed) > 1:
                        header = [all_parsed[0]]
                        rows = all_parsed[1:]
                        total_count = len(rows)
                        if limit is not None:
                            paginated_rows = rows[offset:offset + limit]
                        else:
                            paginated_rows = rows[offset:]
                        parsed = header + paginated_rows
                        if total_rows_count is None:
                            data['total_count'] = total_count
                        else:
                            data['total_count'] = total_rows_count
                    else:
                        total_count = len(all_parsed)
                        if limit is not None:
                            parsed = all_parsed[offset:offset + limit]
                        else:
                            parsed = all_parsed[offset:]
                        if total_rows_count is None:
                            data['total_count'] = total_count
                        else:
                            data['total_count'] = total_rows_count
                except Exception as e:
                    logger.error(f"Ошибка при чтении Excel через Polars: {e}")
                    # Fallback на старый метод
                    try:
                        all_parsed, sheets = self._parse_xlsx(instance.file.path, has_header)
                        data['sheets'] = sheets
                        
                        # Применяем пагинацию
                        if has_header and len(all_parsed) > 1:
                            header = [all_parsed[0]]
                            rows = all_parsed[1:]
                            total_count = len(rows)
                            if limit is not None:
                                paginated_rows = rows[offset:offset + limit]
                            else:
                                paginated_rows = rows[offset:]
                            parsed = header + paginated_rows
                            if total_rows_count is None:
                                data['total_count'] = total_count
                            else:
                                data['total_count'] = total_rows_count
                        else:
                            total_count = len(all_parsed)
                            if limit is not None:
                                parsed = all_parsed[offset:offset + limit]
                            else:
                                parsed = all_parsed[offset:]
                            if total_rows_count is None:
                                data['total_count'] = total_count
                            else:
                                data['total_count'] = total_rows_count
                    except Exception as e2:
                        logger.error(f"Ошибка при чтении Excel через fallback метод: {e2}")
                        raise ValidationError(f"Ошибка чтения Excel файла: {str(e2)}")

            else:
                raise ValidationError(f"Неподдерживаемый тип файла: {instance.file_type}")

            data['parsed'] = parsed
            return Response(data)

        except Exception as exc:
            logger.error(f"Ошибка при чтении файла {instance.id}: {str(exc)}", exc_info=True)
            return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def perform_update(self, serializer):
        instance = serializer.instance
        
        # Убеждаемся, что UUID есть
        if not instance.file_uuid:
            instance.file_uuid = uuid4()
            instance.save(update_fields=['file_uuid'])
        
        file = self.request.FILES.get('file')
        # Безопасный доступ к request.data для избежания ошибок типизации
        request_data = getattr(self.request, 'data', {})
        name = request_data.get('name') if isinstance(request_data, dict) else None
        sheet = request_data.get('sheet') if isinstance(request_data, dict) else None
        file_type = (file.name.split('.')[-1].lower()
                     if file else 
                     (name.split('.')[-1].lower() if name and '.' in name else None))

        # Если есть новый файл и указан лист - обрабатываем листы
        if file and file_type == 'xlsx' and sheet:
            print(f"[DEBUG UPDATE] Обновляем Excel файл с листом: {sheet}")
            suffix = os.path.splitext(file.name)[-1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                for chunk in file.chunks():
                    tmp.write(chunk)
                temp_path = tmp.name
            
            try:
                wb = load_workbook(temp_path, read_only=False)
                print(f"[DEBUG UPDATE] Доступные листы: {wb.sheetnames}")
                
                # Удаляем все листы кроме нужного
                sheets_to_remove = [ws_name for ws_name in wb.sheetnames if ws_name != sheet]
                for ws_name in sheets_to_remove:
                    ws = wb[ws_name]
                    wb.remove(ws)
                
                single_sheet_path = temp_path + "_single.xlsx"
                wb.save(single_sheet_path)
                wb.close()
                print(f"[DEBUG UPDATE] Обновляем файл одним листом: {single_sheet_path}")
                
                # Сначала сохраняем экземпляр без файла
                instance = serializer.save(
                    name=name or serializer.instance.name,
                    original_filename=file.name,
                    file_type=file_type
                )
                
                # Удаляем старый файл если он есть
                if instance.file and hasattr(instance.file, 'path') and os.path.exists(instance.file.path):
                    try:
                        old_path = instance.file.path
                        instance.file.delete(save=False)
                        print(f"[DEBUG UPDATE] Удален старый файл: {old_path}")
                    except Exception as e:
                        print(f"[DEBUG UPDATE] Ошибка при удалении старого файла: {e}")
                
                # Затем обновляем файл отдельно с UUID в имени
                uuid_filename = f"{instance.file_uuid}.bin" if file_type == 'bin' else f"{instance.file_uuid}{os.path.splitext(file.name)[1]}"
                with open(single_sheet_path, 'rb') as f:
                    instance.file.save(uuid_filename, File(f), save=True)
                
                print(f"[DEBUG UPDATE] Файл обновлен в базе: {instance.file.path}")
                
                # Очищаем временные файлы
                try:
                    os.remove(single_sheet_path)
                    os.remove(temp_path)
                except Exception as e:
                    print(f"[DEBUG UPDATE] Ошибка при удалении временных файлов: {e}")
                    
            except Exception as e:
                print(f"[DEBUG UPDATE] Ошибка при обработке листов: {e}")
                # Если не удалось обработать листы, сохраняем как есть
                try:
                    os.remove(temp_path)
                except Exception:
                    pass
                # Сохраняем файл с UUID в имени
                if file:
                    file_ext = os.path.splitext(file.name)[1] or ('.bin' if file_type == 'bin' else '.xlsx')
                    uuid_filename = f"{instance.file_uuid}{file_ext}"
                    with open(temp_path, 'rb') as f:
                        instance.file.save(uuid_filename, File(f), save=True)
                
                instance = serializer.save(
                    name=name or serializer.instance.name,
                    original_filename=(file.name if file else serializer.instance.original_filename),
                    file_type=file_type
                )
        else:
            # Сохраняем файл с UUID в имени если есть новый файл
            if file:
                file_ext = os.path.splitext(file.name)[1] or ('.bin' if file_type == 'bin' else '.xlsx')
                uuid_filename = f"{instance.file_uuid}{file_ext}"
                instance.file.save(uuid_filename, file, save=False)
            
            instance = serializer.save(
                name=name or serializer.instance.name,
                original_filename=(file.name if file else serializer.instance.original_filename),
                file_type=file_type
            )

        columns_info = extract_columns_info(instance)
        instance.columns_info = columns_info
        instance.save(update_fields=['columns_info'])

    def perform_destroy(self, instance):
        # удаляем сам файл
        if instance.file and os.path.isfile(instance.file.path):
            try:
                os.remove(instance.file.path)
            except PermissionError:
                import gc
                gc.collect()
                raise ValidationError("Файл занят другим процессом.")
        instance.delete()

    # -----------------------------
    # вспомогательные методы
    # -----------------------------
    @staticmethod
    def _parse_csv(path, encoding, delimiter, has_header):
        with open(path, 'r', encoding=encoding) as f:
            reader = csv.reader(f, delimiter='\t' if delimiter == '\\t' else delimiter)
            rows = list(reader)
        if not has_header and rows:
            alph = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            headers = [alph[i] if i < len(alph) else f"Col{i}" for i in range(len(rows[0]))]
            rows.insert(0, headers)
        return rows

    @staticmethod
    def _parse_xlsx(path, has_header):
        wb = load_workbook(filename=path, read_only=True)
        try:
            sheet = wb.sheetnames[0]
            ws = wb[sheet]
            data = list(ws.values)
            sheetnames = wb.sheetnames
            if not has_header:
                return data, sheetnames
            headers, *body = data
            return [list(headers), *body], sheetnames
        finally:
            wb.close()

class FileUploadByConnectionView(generics.ListAPIView):
    """
    GET /connection/{id}/files/ — файлы, загруженные к указанному соединению
    """
    serializer_class = FileUploadSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        conn_id = self.kwargs['connection_id']
        # Оптимизация: используем select_related для уменьшения количества запросов к БД
        qs = FileUpload.objects.filter(
            owner=self.request.user, 
            connection_id=conn_id
        ).select_related('connection', 'owner').order_by('-uploaded_at')
        return qs
    
class DatasetRowsAggAPIView(APIView):
    """
    POST  .../datasets/<dataset_id>/rows-agg/
    body = { "fields": {... exactly Chart.params ...} }
    """
    def post(self, request, pk):
        ds = get_object_or_404(Dataset, pk=pk)
        chart_fields = []
        request_data = getattr(request, 'data', {})
        fields_data = request_data.get('fields', {}) if isinstance(request_data, dict) else {}
        for group_key, field_list in (fields_data or {}).items():
            chart_fields.extend(field_list)
        data = get_rows_for_chart(ds, chart_fields)
        return Response(data)

def detect_column_type(values):
    filtered = [v for v in values if v not in (None, '')]
    if not filtered:
        return "string"
    if all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in filtered):
        if all(isinstance(v, int) for v in filtered):
            return "integer"
        return "float"
    try:
        import dateutil.parser
        if all(isinstance(dateutil.parser.parse(str(v)), object) for v in filtered):
            return "date"
    except Exception:
        pass
    # bool
    if all(str(v).lower() in ('true', 'false') for v in filtered):
        return "bool"
    return "string"

def extract_columns_info(instance):
    """
    Оптимизированная функция для извлечения информации о колонках.
    Читает только первые 200 строк вместо всего файла для ускорения.
    Поддерживает бинарные файлы .bin.
    """
    path = instance.file.path
    # Проверяем, является ли файл бинарным
    from src.core.bi_analysis.bi_datasets.binary_storage import is_binary_file, read_from_binary
    
    if is_binary_file(path) or instance.file_type == 'bin':
        try:
            # Читаем только первые 200 строк для определения типов колонок (для производительности)
            columns, rows = read_from_binary(path, row_limit=200)
            if not rows:
                return {'columns': [], 'types': []}
            
            # Определяем типы колонок
            types = []
            sample_rows = rows[:50] if len(rows) > 50 else rows
            for col_idx in range(len(columns)):
                col_values = [row[col_idx] for row in sample_rows if len(row) > col_idx and row[col_idx] is not None]
                types.append(detect_column_type(col_values))
            
            return {'columns': columns, 'types': types}
        except Exception as e:
            logger.error(f"Ошибка чтения бинарного файла для извлечения колонок: {str(e)}")
            return {'columns': [], 'types': []}
    
    if instance.file_type == 'xlsx':
        try:
            wb = openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
            sheet = wb.active
            
            # Читаем только первые 200 строк для определения типов колонок (для производительности)
            rows = []
            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if idx >= 200:
                    break
                rows.append(row)
            
            if not rows:
                return {'columns': [], 'types': []}
            
            headers = rows[0] if rows else []
            columns = list(headers)
            types = []
            
            # Используем только первые 50 строк для определения типов (как было)
            sample_rows = rows[1:51] if len(rows) > 1 else []
            for col_idx in range(len(columns)):
                col_values = [row[col_idx] for row in sample_rows if len(row) > col_idx and row[col_idx] is not None]
                types.append(detect_column_type(col_values))
            
            wb.close()
            return {'columns': columns, 'types': types}
        except Exception:
            return {'columns': [], 'types': []}

    elif instance.file_type == 'csv':
        try:
            # Для CSV читаем построчно, останавливаемся после 200 строк для определения типов (для производительности)
            rows = []
            with open(path, encoding='utf-8') as f:
                reader = csv.reader(f)
                for idx, row in enumerate(reader):
                    if idx >= 200:
                        break
                    rows.append(row)
            
            if not rows:
                return {'columns': [], 'types': []}
            
            headers = rows[0] if rows else []
            columns = list(headers)
            types = []
            
            # Используем только первые 50 строк для определения типов
            sample_rows = rows[1:51] if len(rows) > 1 else []
            for col_idx in range(len(columns)):
                col_values = [row[col_idx] for row in sample_rows if len(row) > col_idx and row[col_idx]]
                types.append(detect_column_type(col_values))
            
            return {'columns': columns, 'types': types}
        except Exception:
            return {'columns': [], 'types': []}
    return {'columns': [], 'types': []}

class FinalizeUploadView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        request_data = getattr(request, 'data', {})
        if not isinstance(request_data, dict):
            request_data = {}
        temp_path         = request_data.get('temp_path')
        name              = request_data.get('name')
        original_filename = request_data.get('original_filename')
        file_type         = request_data.get('file_type')
        connection_id     = request_data.get('connection')
        sheet             = request_data.get('sheet')

        if not all([temp_path, name, original_filename, file_type]):
            return Response(
                {"error": "Необходимы поля temp_path, name, original_filename, file_type"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not os.path.exists(temp_path):
            return Response({"error": "Временный файл не найден"}, status=status.HTTP_404_NOT_FOUND)

        connection_obj = None
        if connection_id:
            try:
                connection_obj = Connection.objects.get(pk=connection_id)
            except Connection.DoesNotExist:
                return Response({"error": "Connection not found"}, status=404)

        upload = FileUpload(
            owner=request.user,
            name=name,
            original_filename=original_filename,
            file_type=file_type,
            connection=connection_obj
        )

        # Сначала сохраняем объект без файла, чтобы сгенерировался UUID
        upload.save()
        
        # UUID должен быть сгенерирован автоматически при сохранении
        if not upload.file_uuid:
            from uuid import uuid4
            upload.file_uuid = uuid4()
            upload.save(update_fields=['file_uuid'])
        
        logger.info(f"Создан FileUpload с UUID: {upload.file_uuid}, ID: {upload.id}")
        
        # Конвертируем файл в бинарный формат .bin перед сохранением
        binary_converted = False
        binary_path = None
        
        # Всегда пытаемся конвертировать в бинарный формат для всех типов файлов
        try:
            from src.core.bi_analysis.bi_datasets.binary_storage import convert_to_binary
            
            # Проверяем, установлен ли Polars
            polars_available = False
            try:
                import polars as pl
                polars_available = True
                logger.info("Polars установлен, конвертация в .bin возможна")
            except ImportError:
                logger.error("Polars не установлен! Файлы не будут конвертироваться в .bin формат")
                binary_converted = False
                binary_path = None
            
            if polars_available:  # Если Polars установлен, продолжаем
                # Конвертируем напрямую из исходного файла, Polars сам прочитает нужный лист
                # Не создаем промежуточный файл, чтобы избежать повреждения данных
                if temp_path and file_type and os.path.exists(temp_path):
                    temp_binary_path = temp_path + ".bin"
                    logger.info(f"[CONVERT] Начинаем конвертацию файла {temp_path} в бинарный формат {temp_binary_path}, тип: {file_type}, лист: {sheet}")
                    
                    conversion_result = convert_to_binary(temp_path, temp_binary_path, sheet_name=sheet, file_type=file_type)
                    logger.info(f"[CONVERT] Результат конвертации: {conversion_result}")
                    
                    if conversion_result:
                        if os.path.exists(temp_binary_path):
                            binary_path = temp_binary_path
                            binary_converted = True
                            file_size = os.path.getsize(binary_path)
                            logger.info(f"[CONVERT] Файл успешно конвертирован в бинарный формат: {binary_path}, размер: {file_size} байт")
                        else:
                            logger.error(f"[CONVERT] Бинарный файл не был создан: {temp_binary_path}")
                            binary_converted = False
                    else:
                        logger.warning(f"[CONVERT] Не удалось конвертировать файл {temp_path} в бинарный формат, используется оригинальный")
                        binary_converted = False
                else:
                    logger.warning(f"[CONVERT] Не удалось найти исходный файл для конвертации: temp_path={temp_path}, file_type={file_type}, exists={os.path.exists(temp_path) if temp_path else False}")
                    binary_converted = False
        except Exception as e:
            logger.error(f"[CONVERT] Ошибка при конвертации в бинарный формат: {str(e)}", exc_info=True)
            binary_converted = False
        
        # Сохраняем файл в Django (бинарный или оригинальный)
        # UUID уже должен быть сгенерирован при первом сохранении
        if not upload.file_uuid:
            upload.file_uuid = uuid4()
            upload.save(update_fields=['file_uuid'])
        
        logger.info(f"[SAVE] Проверка перед сохранением: binary_converted={binary_converted}, binary_path={binary_path}, exists={os.path.exists(binary_path) if binary_path else False}")
        
        if binary_converted and binary_path and os.path.exists(binary_path):
            # Сохраняем бинарный файл с именем по UUID
            # Имя файла = UUID.bin
            binary_filename = f"{upload.file_uuid}.bin"
            
            logger.info(f"[SAVE] Сохраняем бинарный файл с UUID: {binary_filename}, путь: {binary_path}")
            try:
                with open(binary_path, 'rb') as f:
                    upload.file.save(binary_filename, File(f), save=True)
                upload.file_type = 'bin'
                upload.save(update_fields=['file', 'file_type'])
                logger.info(f"[SAVE] Бинарный файл успешно сохранен в Django, file_type установлен в 'bin'")
            except Exception as e:
                logger.error(f"[SAVE] Ошибка при сохранении бинарного файла в Django: {str(e)}", exc_info=True)
                # Fallback на оригинальный файл
                binary_converted = False
            
            # Проверяем, что файл действительно сохранен
            if upload.file and hasattr(upload.file, 'path'):
                saved_path = upload.file.path
                if os.path.exists(saved_path):
                    logger.info(f"Бинарный файл успешно сохранен: {saved_path}, размер: {os.path.getsize(saved_path)} байт")
                else:
                    logger.error(f"Бинарный файл не найден после сохранения: {saved_path}")
            
            # Удаляем временный бинарный файл
            try:
                if binary_path and os.path.exists(binary_path):
                    os.remove(binary_path)
                    logger.info(f"Временный бинарный файл удален: {binary_path}")
            except Exception as e:
                logger.warning(f"Не удалось удалить временный бинарный файл: {e}")
        else:
            # Сохраняем оригинальный файл (fallback) - тоже с UUID
            logger.warning(f"[SAVE] Используется оригинальный файл вместо бинарного. binary_converted={binary_converted}, binary_path={binary_path}, polars_available={polars_available if 'polars_available' in locals() else 'unknown'}")
            
            # Определяем расширение оригинального файла
            if original_filename:
                file_ext = os.path.splitext(original_filename)[1] or '.bin'
            else:
                file_ext = '.bin' if file_type == 'bin' else ('.xlsx' if file_type == 'xlsx' else '.csv')
            
            # Имя файла = UUID + расширение
            uuid_filename = f"{upload.file_uuid}{file_ext}"
            
            if temp_path and os.path.exists(temp_path):
                with open(temp_path, 'rb') as f:
                    upload.file.save(uuid_filename, File(f), save=True)

        upload.columns_info = extract_columns_info(upload)
        upload.save(update_fields=['columns_info'])

        # Удаляем временные файлы
        try:
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)
        except OSError:
            pass
        

        serializer = FileUploadSerializer(upload)
        response_data = serializer.data
        logger.info(f"[FINALIZE] Файл успешно создан: ID={upload.id}, UUID={upload.file_uuid}, name={upload.name}, owner={upload.owner.id}")
        return Response(response_data, status=status.HTTP_201_CREATED)

# ==============================================================================
# XLSX helper endpoints
# ==============================================================================

class XlsxSheetListView(APIView):
    """
    POST /xlsx/sheets/   — получить список листов в загруженном .xlsx-файле
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        file = request.FILES.get('file')
        if not file or not file.name.endswith('.xlsx'):
            raise ValidationError("Ожидался .xlsx файл")

        try:
            wb = load_workbook(filename=file, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return Response({"filename": file.name, "sheets": sheets})
        except Exception as exc:
            return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

class XlsxTempPreviewView(APIView):
    """
    POST /xlsx/preview/  — предпросмотр содержимого временного .xlsx-файла
    Использует Celery для асинхронной обработки больших файлов с polars, чанкингом и векторизацией.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        request_data = getattr(request, 'data', {})
        if not isinstance(request_data, dict):
            request_data = {}
        temp_path     = request_data.get('temp_path')
        has_header    = request_data.get('has_header', 'true').lower() == 'true' if isinstance(request_data.get('has_header'), str) else True
        sheet_name    = request_data.get('sheet_name')
        # Лимиты убраны - если row_limit не указан, загружаем все данные
        row_limit = request_data.get('row_limit')
        if row_limit is not None:
            row_limit = int(row_limit)
        use_async     = request_data.get('async', 'false').lower() == 'true' if isinstance(request_data.get('async'), str) else False

        if not temp_path or not os.path.exists(temp_path):
            return Response({"error": "Временный файл не найден"}, status=status.HTTP_404_NOT_FOUND)

        # Определяем размер файла для выбора стратегии
        file_size = os.path.getsize(temp_path)
        # Для файлов больше 5MB или при явном запросе используем асинхронную обработку
        # Если row_limit не указан (None), считаем что это большой файл и используем асинхронную обработку
        should_use_async = use_async or file_size > 5 * 1024 * 1024 or (row_limit is not None and row_limit > 500) or row_limit is None

        if should_use_async:
            from src.core.bi_analysis.tasks import process_file_preview
            task = process_file_preview.delay(temp_path, sheet_name, row_limit, has_header)
            return Response({
                "task_id": task.id,
                "status": "processing",
                "message": "Предпросмотр обрабатывается асинхронно"
            }, status=202)

        # Синхронная обработка для маленьких файлов
        try:
            # Пробуем использовать polars для быстрой обработки
            try:
                import polars as pl
                
                file_ext = os.path.splitext(temp_path)[1].lower()
                
                if file_ext in ('.xlsx', '.xls'):
                    if sheet_name:
                        df = pl.read_excel(temp_path, sheet_name=sheet_name)
                    else:
                        df = pl.read_excel(temp_path, sheet_index=0)
                    
                    if row_limit and row_limit > 0:
                        df = df.head(row_limit)
                    
                    columns = df.columns
                    rows_list = df.to_numpy().tolist()
                    
                    if has_header and rows_list:
                        parsed = [list(columns), *rows_list]
                    else:
                        parsed = rows_list
                    
                    return Response({"parsed": parsed})
                    
                elif file_ext in ('.csv', '.txt'):
                    try:
                        df = pl.read_csv(temp_path, encoding='utf8', try_parse_dates=True)
                    except:
                        df = pl.read_csv(temp_path, encoding='cp1251', try_parse_dates=True)
                    
                    if row_limit and row_limit > 0:
                        df = df.head(row_limit)
                    
                    columns = df.columns
                    rows_list = df.to_numpy().tolist()
                    
                    if has_header and rows_list:
                        parsed = [list(columns), *rows_list]
                    else:
                        parsed = rows_list
                    
                    return Response({"parsed": parsed})
            except ImportError:
                # Fallback на openpyxl если polars не установлен
                pass

            # Fallback на openpyxl
            wb = load_workbook(filename=temp_path, read_only=True, data_only=True)
            try:
                ws = wb.active
            except IndexError:
                wb.close()
                return Response({"parsed": []})

            rows = []
            for row in ws.iter_rows(values_only=True):
                normalized = [("" if cell is None else str(cell)) for cell in row]
                rows.append(normalized)
                if len(rows) >= row_limit:
                    break
            wb.close()

            if not rows:
                return Response({"parsed": []})

            if has_header:
                header, *body = rows
                parsed = [list(header), *body]
            else:
                parsed = rows

            return Response({"parsed": parsed})
        except Exception as exc:
            return Response({"error": f"Ошибка при чтении файла: {exc}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class FilePreviewTaskStatusView(APIView):
    """
    GET /xlsx/preview/task-status/  — получение статуса задачи предпросмотра файла
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """
        Получает статус и результат асинхронной задачи предпросмотра файла.
        """
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({"detail": "Не указан task_id"}, status=400)
        
        try:
            from celery.result import AsyncResult
            from src.config.celery import celery_app
            
            task = AsyncResult(task_id, app=celery_app)
            
            if task.state == 'PENDING':
                response = {
                    'task_id': task_id,
                    'status': 'pending',
                    'message': 'Задача ожидает выполнения'
                }
            elif task.state == 'PROGRESS':
                response = {
                    'task_id': task_id,
                    'status': 'processing',
                    'message': task.info.get('message', 'Задача выполняется') if isinstance(task.info, dict) else 'Задача выполняется',
                    'progress': task.info.get('progress', 0) if isinstance(task.info, dict) else None
                }
            elif task.state == 'SUCCESS':
                result = task.result
                response = {
                    'task_id': task_id,
                    'status': 'success',
                    'result': result
                }
            elif task.state == 'FAILURE':
                response = {
                    'task_id': task_id,
                    'status': 'failure',
                    'error': str(task.info) if task.info else 'Неизвестная ошибка'
                }
            else:
                response = {
                    'task_id': task_id,
                    'status': task.state.lower(),
                    'message': f'Статус задачи: {task.state}'
                }
            
            return Response(response, status=200)
            
        except Exception as e:
            return Response({"detail": f"Ошибка получения статуса задачи: {str(e)}"}, status=500)

# ==============================================================================
# Field values endpoint
# ==============================================================================

class DatasetFieldValuesView(APIView):
    """
    GET /bi_analysis/bi_datasets/{pk}/field-values/{field_id}/
    Получить уникальные значения для конкретного поля датасета
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk, field_id):
        dataset = Dataset.objects.filter(pk=pk, owner=request.user).first()
        
        if not dataset:
            dataset_exists = Dataset.objects.filter(pk=pk).exists()
            
            if dataset_exists:
                return Response({"detail": "Dataset exists but doesn't belong to current user"}, status=404)
            else:
                return Response({"detail": "Dataset not found"}, status=404)

        field = DataSetField.objects.filter(pk=field_id, dataset=dataset).first()
        
        if not field:
            field_exists = DataSetField.objects.filter(pk=field_id).exists()
            
            if field_exists:
                return Response({"detail": "Field exists but doesn't belong to this dataset"}, status=404)
            else:
                return Response({"detail": "Field not found"}, status=404)

        try:
            base_query, display_columns = build_dataset_query(
                dataset,
                select_fields=[field.name],
                limit=None,
                offset=None,
            )
            # Базовый запрос возвращает одну колонку out_0 при select_fields
            col_ident = sql.Identifier('out_0')

            distinct_query = sql.SQL(
                'SELECT DISTINCT {col} '
                'FROM ({base}) AS sub '
                'WHERE {col} IS NOT NULL '
                'ORDER BY {col} '
                'LIMIT 1000'
            ).format(
                col=col_ident,
                base=base_query,
            )

            with connection.cursor() as cursor:
                # Получаем уникальные значения поля
                cursor.execute(distinct_query)
                rows = cursor.fetchall()
                values = [str(row[0]) for row in rows if row[0] is not None]
        except ProgrammingError as e:
            return Response(
                {"detail": f"Ошибка выполнения SQL запроса для поля {field.source_column}: {str(e)}"},
                status=404,
            )
        except Exception as e:
            return Response({"detail": f"Database error: {str(e)}"}, status=500)

        return Response({
            "field_id": field_id,
            "field_name": field.name,
            "field_column": field.source_column,
            "values": values,
            "count": len(values)
        })

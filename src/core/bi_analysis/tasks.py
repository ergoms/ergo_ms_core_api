"""
Celery задачи для асинхронной обработки предпросмотра датасетов.
Использует polars для быстрой обработки файлов с поддержкой чанкинга и параллельной обработки.
"""

import logging
import os
from typing import List, Dict, Any, Optional, Tuple, Callable
from concurrent.futures import ThreadPoolExecutor, as_completed
from celery import shared_task
from django.db import connection
from psycopg2 import sql
from rest_framework.exceptions import ValidationError

from src.core.bi_analysis.bi_datasets.models import FileUpload
from src.core.bi_analysis.config import (
    get_compute_device, 
    CHUNK_SIZE, 
    MAX_WORKERS
)
from src.core.bi_analysis.services.services import introspect_columns

logger = logging.getLogger(__name__)


def _read_file_with_polars(file_upload_id: int, sheet_name: Optional[str] = None, 
                           row_limit: Optional[int] = None, 
                           progress_callback=None) -> Tuple[List[str], List[List[Any]]]:
    """
    Читает файл используя polars для максимальной производительности с векторизацией.
    
    Args:
        file_upload_id: ID загруженного файла
        sheet_name: Имя листа для Excel файлов
        row_limit: Лимит строк для чтения
        progress_callback: Функция для обновления прогресса (progress: float)
        
    Returns:
        Tuple[columns, rows] - список колонок и список строк
    """
    try:
        import polars as pl
    except ImportError:
        logger.warning("Polars не установлен, используется fallback метод")
        from src.core.bi_analysis.services.services import read_file_to_dataframe
        table = read_file_to_dataframe(file_upload_id, sheet_name, row_limit)
        return table.columns, table.rows
    
    try:
        upload = FileUpload.objects.get(pk=file_upload_id)
    except FileUpload.DoesNotExist:
        logger.error(f"FileUpload с id={file_upload_id} не найден")
        raise ValueError(f"FileUpload с id={file_upload_id} не найден")
    
    if not upload.file:
        logger.error(f"FileUpload {file_upload_id} не имеет файла")
        raise ValueError(f"FileUpload {file_upload_id} не имеет файла")
    
    path = upload.file.path
    
    # Проверяем, является ли файл бинарным
    from src.core.bi_analysis.bi_datasets.binary_storage import is_binary_file, read_from_binary
    
    if is_binary_file(path) or upload.file_type == 'bin':
        # Читаем из бинарного файла
        if progress_callback:
            progress_callback(0.1)
        try:
            columns, rows_list = read_from_binary(path, row_limit=row_limit)
            if progress_callback:
                progress_callback(1.0)
            return columns, rows_list
        except Exception as e:
            logger.error(f"Ошибка чтения бинарного файла {file_upload_id}: {str(e)}")
            raise
    
    if progress_callback:
        progress_callback(0.1)
    
    try:
        if upload.file_type == 'xlsx':
            # Читаем Excel файл с polars (векторизованная операция)
            if sheet_name:
                df = pl.read_excel(path, sheet_name=sheet_name)
            else:
                # Берем первый лист
                df = pl.read_excel(path, sheet_index=0)
            
            if progress_callback:
                progress_callback(0.5)
            
            # Используем векторизованные операции polars для ограничения
            # Если row_limit не указан (None), читаем все данные
            if row_limit is not None and row_limit > 0:
                df = df.head(row_limit)
            
            # Векторизованная конвертация в списки
            columns = df.columns
            # Используем to_numpy для быстрой конвертации (векторизация)
            rows_list = df.to_numpy().tolist()
            
            if progress_callback:
                progress_callback(1.0)
            
            return list(columns), rows_list
            
        elif upload.file_type in ('csv', 'txt'):
            # Определяем кодировку автоматически
            try:
                df = pl.read_csv(path, encoding='utf8', try_parse_dates=True)
            except:
                # Fallback на cp1251
                df = pl.read_csv(path, encoding='cp1251', try_parse_dates=True)
            
            if progress_callback:
                progress_callback(0.5)
            
            # Векторизованная операция ограничения
            # Если row_limit не указан (None), читаем все данные
            if row_limit is not None and row_limit > 0:
                df = df.head(row_limit)
            
            columns = df.columns
            # Векторизованная конвертация
            rows_list = df.to_numpy().tolist()
            
            if progress_callback:
                progress_callback(1.0)
            
            return list(columns), rows_list
        else:
            raise ValidationError(f"Неподдерживаемый тип файла: {upload.file_type}")
            
    except Exception as e:
        logger.error(f"Ошибка чтения файла {file_upload_id} с polars: {str(e)}")
        # Fallback на старый метод
        from src.core.bi_analysis.services.services import read_file_to_dataframe
        table = read_file_to_dataframe(file_upload_id, sheet_name, row_limit)
        return table.columns, table.rows


def _read_file_chunked(file_upload_id: int, sheet_name: Optional[str] = None,
                      row_limit: Optional[int] = None,
                      progress_callback=None) -> Tuple[List[str], List[List[Any]]]:
    """
    Читает файл по частям (чанками) для больших файлов с векторизацией.
    Использует polars с streaming для обработки больших файлов.
    """
    try:
        import polars as pl
    except ImportError:
        return _read_file_with_polars(file_upload_id, sheet_name, row_limit, progress_callback)
    
    try:
        upload = FileUpload.objects.get(pk=file_upload_id)
    except FileUpload.DoesNotExist:
        logger.error(f"FileUpload с id={file_upload_id} не найден")
        raise ValueError(f"FileUpload с id={file_upload_id} не найден")
    
    if not upload.file:
        logger.error(f"FileUpload {file_upload_id} не имеет файла")
        raise ValueError(f"FileUpload {file_upload_id} не имеет файла")
    
    path = upload.file.path
    
    # Определяем размер файла для оценки прогресса
    file_size = os.path.getsize(path)
    estimated_rows = file_size // 100  # Примерная оценка
    
    try:
        if upload.file_type == 'xlsx':
            # Для Excel используем чанкинг через polars
            # Polars может читать Excel по частям через read_excel с параметрами
            if sheet_name:
                df = pl.read_excel(path, sheet_name=sheet_name)
            else:
                df = pl.read_excel(path, sheet_index=0)
            
            if progress_callback:
                progress_callback(0.3)
            
            # Векторизованная обработка: применяем limit через head
            if row_limit and row_limit > 0:
                df = df.head(row_limit)
            
            if progress_callback:
                progress_callback(0.7)
            
            columns = df.columns
            # Векторизованная конвертация
            rows_list = df.to_numpy().tolist()
            
            if progress_callback:
                progress_callback(1.0)
            
            return list(columns), rows_list
            
        elif upload.file_type in ('csv', 'txt'):
            # Используем streaming для CSV с векторизацией
            columns = None
            all_rows = []
            rows_read = 0
            
            try:
                # Пробуем UTF-8
                reader = pl.scan_csv(path, encoding='utf8', try_parse_dates=True)
            except:
                # Fallback на cp1251
                reader = pl.scan_csv(path, encoding='cp1251', try_parse_dates=True)
            
            # Получаем схему
            schema = reader.schema
            columns = list(schema.keys())
            
            if progress_callback:
                progress_callback(0.1)
            
            # Читаем по чанкам с векторизацией
            chunk_count = 0
            for chunk in reader.iter_slices(n_rows=CHUNK_SIZE):
                chunk_df = chunk.collect()
                # Векторизованная конвертация через to_numpy
                chunk_rows_list = chunk_df.to_numpy().tolist()
                
                all_rows.extend(chunk_rows_list)
                rows_read += len(chunk_rows_list)
                chunk_count += 1
                
                # Обновляем прогресс
                if progress_callback and estimated_rows > 0:
                    progress = min(0.9, 0.1 + (rows_read / estimated_rows) * 0.8)
                    progress_callback(progress)
                
                if row_limit and row_limit > 0 and rows_read >= row_limit:
                    all_rows = all_rows[:row_limit]
                    break
            
            if progress_callback:
                progress_callback(1.0)
            
            return columns, all_rows
        else:
            return _read_file_with_polars(file_upload_id, sheet_name, row_limit, progress_callback)
            
    except Exception as e:
        logger.error(f"Ошибка чтения файла {file_upload_id} по чанкам: {str(e)}")
        return _read_file_with_polars(file_upload_id, sheet_name, row_limit, progress_callback)


def _read_files_parallel(file_data_list: List[Dict[str, Any]]) -> List[Tuple[List[str], List[List[Any]]]]:
    """
    Читает несколько файлов параллельно используя ThreadPoolExecutor.
    
    Args:
        file_data_list: Список словарей с ключами file_id, sheet_name, row_limit
        
    Returns:
        Список кортежей (columns, rows) для каждого файла
    """
    results = [None] * len(file_data_list)
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(
                _read_file_with_polars,
                file_data['file_id'],
                file_data.get('sheet_name'),
                file_data.get('row_limit')
            ): idx
            for idx, file_data in enumerate(file_data_list)
        }
        
        for future in as_completed(futures):
            idx = futures[future]
            try:
                results[idx] = future.result()
            except Exception as e:
                logger.error(f"Ошибка чтения файла {file_data_list[idx]}: {str(e)}")
                results[idx] = ([], [])
    
    return results


def _read_files_parallel_with_progress(file_data_list: List[Dict[str, Any]], 
                                      progress_callback: Optional[Callable[[int, float], None]] = None) -> List[Tuple[List[str], List[List[Any]]]]:
    """
    Читает несколько файлов параллельно используя ThreadPoolExecutor с поддержкой прогресса.
    Использует чанкинг для больших файлов.
    
    Args:
        file_data_list: Список словарей с ключами file_id, sheet_name, row_limit
        progress_callback: Функция для обновления прогресса (file_idx: int, progress: float)
        
    Returns:
        Список кортежей (columns, rows) для каждого файла
    """
    results = [None] * len(file_data_list)
    
    def read_file_with_progress(file_data, file_idx):
        """Читает файл с обновлением прогресса"""
        def local_progress_callback(progress):
            if progress_callback:
                progress_callback(file_idx, progress)
        
        # Определяем размер файла для выбора стратегии
        try:
            try:
                upload = FileUpload.objects.get(pk=file_data['file_id'])
            except FileUpload.DoesNotExist:
                logger.error(f"FileUpload с id={file_data['file_id']} не найден")
                return ([], [])
            
            if not upload.file:
                logger.error(f"FileUpload {file_data['file_id']} не имеет файла")
                return ([], [])
            
            file_size = os.path.getsize(upload.file.path)
            # Если файл больше 10MB, используем чанкинг
            use_chunked = file_size > 10 * 1024 * 1024
            
            if use_chunked:
                return _read_file_chunked(
                    file_data['file_id'],
                    file_data.get('sheet_name'),
                    file_data.get('row_limit'),
                    local_progress_callback
                )
            else:
                return _read_file_with_polars(
                    file_data['file_id'],
                    file_data.get('sheet_name'),
                    file_data.get('row_limit'),
                    local_progress_callback
                )
        except Exception as e:
            logger.error(f"Ошибка чтения файла {file_data}: {str(e)}")
            return ([], [])
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(read_file_with_progress, file_data, idx): idx
            for idx, file_data in enumerate(file_data_list)
        }
        
        for future in as_completed(futures):
            idx = futures[future]
            try:
                results[idx] = future.result()
            except Exception as e:
                logger.error(f"Ошибка чтения файла {file_data_list[idx]}: {str(e)}")
                results[idx] = ([], [])
    
    return results


@shared_task(bind=True, name='src.core.bi_analysis.tasks.process_dataset_preview')
def process_dataset_preview(self, dataset_id: int, limit: Optional[int] = None):
    """
    Асинхронная задача для обработки предпросмотра датасета.
    Использует polars, векторизацию и параллельную обработку для ускорения.
    Обновляет прогресс через self.update_state.
    """
    try:
        from src.core.bi_analysis.bi_datasets.models import Dataset
        from src.core.bi_analysis.services.services import build_dataset_query
        
        self.update_state(state='PROGRESS', meta={'progress': 0.1, 'message': 'Загрузка датасета'})
        
        dataset = Dataset.objects.get(pk=dataset_id)
        
        self.update_state(state='PROGRESS', meta={'progress': 0.3, 'message': 'Обработка запроса'})
        
        # Проверяем, есть ли таблицы в датасете
        if not dataset.tables.exists():
            raise ValueError("Таблицы не найдены в датасете. Добавьте главную таблицу для предпросмотра.")
        
        # Используем новую логику с динамическим SQL, которая читает файлы напрямую через polars и бинарные файлы
        query = build_dataset_query(dataset, limit=limit)
        
        self.update_state(state='PROGRESS', meta={'progress': 0.5, 'message': 'Выполнение SQL запроса'})
        
        with connection.cursor() as cursor:
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
        
        self.update_state(state='PROGRESS', meta={'progress': 1.0, 'message': 'Завершено'})
        
        return {
            'columns': columns,
            'rows': rows
        }
        
    except Exception as e:
        logger.error(f"Ошибка обработки предпросмотра датасета {dataset_id}: {str(e)}")
        raise


@shared_task(bind=True, name='src.core.bi_analysis.tasks.process_draft_preview')
def process_draft_preview(self, draft_data: Dict[str, Any]):
    """
    Асинхронная задача для обработки предпросмотра черновика датасета.
    Использует polars, векторизацию, чанкинг и параллельную обработку файлов.
    Обновляет прогресс через self.update_state.
    """
    try:
        from src.core.bi_analysis.services.services import dataframe_to_sql_values
        from psycopg2 import sql
        
        self.update_state(state='PROGRESS', meta={'progress': 0.05, 'message': 'Инициализация'})
        
        connection_id = draft_data.get('connection_id')
        main_table = draft_data.get('mainTable')
        
        # Проверяем наличие главной таблицы
        if not main_table or (isinstance(main_table, dict) and not main_table):
            raise ValidationError("Не указана главная таблица")
        
        from django.conf import settings
        
        joined_tables = draft_data.get('joinedTables', [])
        # Лимиты убраны - если limit не указан, загружаем все данные
        limit = draft_data.get('limit')
        if limit is not None:
            limit = int(limit)
        
        def get_sheet_name(table_dict):
            if not isinstance(table_dict, dict):
                return None
            sheet = (
                table_dict.get('sheet_name') or
                table_dict.get('sheetName') or
                table_dict.get('sheet')
            )
            if sheet:
                return sheet
            nested_file = table_dict.get('file_upload') or {}
            if isinstance(nested_file, dict):
                return nested_file.get('sheet_name') or nested_file.get('sheet')
            return None
        
        # Определяем, является ли главная таблица файловым источником
        is_main_file = 'file_id' in main_table and main_table.get('file_id')
        main_alias = 'a'
        
        # Подготавливаем данные для параллельного чтения файлов
        files_to_read = []
        if is_main_file:
            files_to_read.append({
                'file_id': main_table['file_id'],
                'sheet_name': get_sheet_name(main_table),
                'row_limit': limit * 2  # Читаем больше для JOIN'ов
            })
        
        for jt in joined_tables:
            if 'file_id' in jt and jt.get('file_id'):
                files_to_read.append({
                    'file_id': jt['file_id'],
                    'sheet_name': get_sheet_name(jt),
                    'row_limit': limit * 2
                })
        
        self.update_state(state='PROGRESS', meta={'progress': 0.15, 'message': f'Чтение {len(files_to_read)} файлов'})
        
        # Читаем файлы параллельно с прогрессом
        if files_to_read:
            # Создаем callback для прогресса
            def progress_callback(file_idx, progress):
                total_progress = 0.15 + (file_idx / len(files_to_read)) * 0.5 + (progress * 0.5 / len(files_to_read))
                self.update_state(state='PROGRESS', meta={
                    'progress': total_progress,
                    'message': f'Обработка файла {file_idx + 1}/{len(files_to_read)}: {int(progress * 100)}%'
                })
            
            # Модифицируем _read_files_parallel для поддержки прогресса
            file_results = _read_files_parallel_with_progress(files_to_read, progress_callback)
            file_index = 0
        
        # Обрабатываем главную таблицу
        if is_main_file:
            main_cols, main_rows = file_results[file_index]
            file_index += 1
            
            # Создаем TableData для совместимости
            from src.core.bi_analysis.services.services import TableData
            main_table_data = TableData(main_cols, main_rows)
            main_from, _ = dataframe_to_sql_values(main_table_data, main_alias)
        else:
            # Для БД источников
            table_name = main_table.get('table_name')
            if not table_name:
                raise ValidationError("Не указано имя таблицы для БД источника")
            
            if '.' in table_name:
                schema, table = table_name.split('.', 1)
                main_from = sql.SQL('{}.{} AS {}').format(
                    sql.Identifier(schema),
                    sql.Identifier(table),
                    sql.Identifier(main_alias)
                )
            else:
                main_from = sql.SQL('{} AS {}').format(
                    sql.Identifier(table_name),
                    sql.Identifier(main_alias)
                )
            
            main_cols = introspect_columns(table_name)
        
        # Формируем SELECT
        select_parts = [sql.SQL('{}.{} AS {}').format(
            sql.Identifier(main_alias),
            sql.Identifier(col),
            sql.Identifier(col)
        ) for col in main_cols]
        
        # Обрабатываем JOIN'ы
        join_clauses = []
        alias_idx = ord('b')
        all_cols = set(main_cols)
        file_index = 1 if is_main_file else 0
        
        for jt in joined_tables:
            tbl_alias = chr(alias_idx)
            alias_idx += 1
            
            is_join_file = 'file_id' in jt and jt.get('file_id')
            
            if is_join_file:
                join_cols, join_rows = file_results[file_index]
                file_index += 1
                
                from src.core.bi_analysis.services.services import TableData
                join_table_data = TableData(join_cols, join_rows)
                join_from, _ = dataframe_to_sql_values(join_table_data, tbl_alias)
            else:
                table_name = jt.get('table_name')
                if not table_name:
                    raise ValidationError("Не указано имя таблицы для JOIN БД источника")
                
                if '.' in table_name:
                    schema, table = table_name.split('.', 1)
                    join_from = sql.SQL('{}.{} AS {}').format(
                        sql.Identifier(schema),
                        sql.Identifier(table),
                        sql.Identifier(tbl_alias)
                    )
                else:
                    join_from = sql.SQL('{} AS {}').format(
                        sql.Identifier(table_name),
                        sql.Identifier(tbl_alias)
                    )
                
                join_cols = introspect_columns(table_name)
            
            # Добавляем колонки из JOIN таблицы
            for col in join_cols:
                if col not in all_cols:
                    select_parts.append(sql.SQL('{}.{} AS {}').format(
                        sql.Identifier(tbl_alias),
                        sql.Identifier(col),
                        sql.Identifier(col)
                    ))
                    all_cols.add(col)
            
            # Формируем условие JOIN
            join_type = jt.get('joinType', 'LEFT JOIN').strip().upper()
            if 'JOIN' not in join_type:
                join_type = f'{join_type} JOIN'
            
            lines = jt.get('lines') or []
            if not lines or not lines[0].get('left') or not lines[0].get('right'):
                raise ValidationError("Не указаны поля для join'а")
            
            left_col = lines[0]['left']
            right_col = lines[0]['right']
            
            join_condition = sql.SQL('{}.{} = {}.{}').format(
                sql.Identifier(main_alias),
                sql.Identifier(left_col),
                sql.Identifier(tbl_alias),
                sql.Identifier(right_col)
            )
            
            join_clause = sql.SQL('{} {} ON {}').format(
                sql.SQL(join_type),
                join_from,
                join_condition
            )
            join_clauses.append(join_clause)
        
        # Собираем итоговый запрос
        query_parts = [
            sql.SQL('SELECT {}').format(sql.SQL(', ').join(select_parts)),
            sql.SQL('FROM {}').format(main_from)
        ]
        
        query_parts.extend(join_clauses)
        query_parts.append(sql.SQL('LIMIT {}').format(sql.Literal(limit)))
        
        final_query = sql.SQL(' ').join(query_parts)
        
        # Выполняем запрос
        with connection.cursor() as cursor:
            cursor.execute(final_query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
        
        self.update_state(state='PROGRESS', meta={'progress': 0.9, 'message': 'Формирование результата'})
        
        result = {
            'columns': columns,
            'rows': rows
        }
        
        self.update_state(state='PROGRESS', meta={'progress': 1.0, 'message': 'Завершено'})
        
        return result
        
    except Exception as e:
        logger.error(f"Ошибка обработки предпросмотра черновика: {str(e)}")
        raise


@shared_task(bind=True, name='src.core.bi_analysis.tasks.process_file_preview')
def process_file_preview(self, temp_path: str, sheet_name: Optional[str] = None, 
                        row_limit: int = 1000000, has_header: bool = True):
    """
    Асинхронная задача для обработки предпросмотра временного файла.
    Использует polars с векторизацией и чанкингом для максимальной производительности.
    Обновляет прогресс через self.update_state.
    """
    try:
        import polars as pl
        
        self.update_state(state='PROGRESS', meta={'progress': 0.1, 'message': 'Определение типа файла'})
        
        # Определяем тип файла по расширению
        file_ext = os.path.splitext(temp_path)[1].lower()
        
        if file_ext in ('.xlsx', '.xls'):
            # Читаем Excel файл с polars (векторизованная операция)
            self.update_state(state='PROGRESS', meta={'progress': 0.3, 'message': 'Чтение Excel файла'})
            
            if sheet_name:
                df = pl.read_excel(temp_path, sheet_name=sheet_name)
            else:
                # Берем первый лист
                df = pl.read_excel(temp_path, sheet_index=0)
            
            self.update_state(state='PROGRESS', meta={'progress': 0.6, 'message': 'Обработка данных'})
            
            # Векторизованная операция ограничения
            # Если row_limit не указан (None), читаем все данные
            if row_limit is not None and row_limit > 0:
                df = df.head(row_limit)
            
            # Векторизованная конвертация в списки
            columns = df.columns
            rows_list = df.to_numpy().tolist()
            
            self.update_state(state='PROGRESS', meta={'progress': 0.9, 'message': 'Формирование результата'})
            
            # Обрабатываем заголовок
            if has_header and rows_list:
                parsed = [list(columns), *rows_list]
            else:
                parsed = rows_list
            
            self.update_state(state='PROGRESS', meta={'progress': 1.0, 'message': 'Завершено'})
            
            return {'parsed': parsed}
            
        elif file_ext in ('.csv', '.txt'):
            # Читаем CSV файл с polars
            self.update_state(state='PROGRESS', meta={'progress': 0.3, 'message': 'Чтение CSV файла'})
            
            try:
                df = pl.read_csv(temp_path, encoding='utf8', try_parse_dates=True)
            except:
                # Fallback на cp1251
                df = pl.read_csv(temp_path, encoding='cp1251', try_parse_dates=True)
            
            self.update_state(state='PROGRESS', meta={'progress': 0.6, 'message': 'Обработка данных'})
            
            # Векторизованная операция ограничения
            # Если row_limit не указан (None), читаем все данные
            if row_limit is not None and row_limit > 0:
                df = df.head(row_limit)
            
            columns = df.columns
            rows_list = df.to_numpy().tolist()
            
            self.update_state(state='PROGRESS', meta={'progress': 0.9, 'message': 'Формирование результата'})
            
            # Обрабатываем заголовок
            if has_header and rows_list:
                parsed = [list(columns), *rows_list]
            else:
                parsed = rows_list
            
            self.update_state(state='PROGRESS', meta={'progress': 1.0, 'message': 'Завершено'})
            
            return {'parsed': parsed}
        else:
            raise ValidationError(f"Неподдерживаемый тип файла: {file_ext}")
            
    except ImportError:
        # Fallback на openpyxl для Excel
        logger.warning("Polars не установлен, используется fallback метод")
        from openpyxl import load_workbook
        
        self.update_state(state='PROGRESS', meta={'progress': 0.3, 'message': 'Чтение файла (fallback)'})
        
        wb = load_workbook(filename=temp_path, read_only=True, data_only=True)
        try:
            ws = wb.active
        except IndexError:
            wb.close()
            return {'parsed': []}
        
        rows = []
        for row in ws.iter_rows(values_only=True):
            normalized = [("" if cell is None else str(cell)) for cell in row]
            rows.append(normalized)
            if len(rows) >= row_limit:
                break
        wb.close()
        
        if not rows:
            return {'parsed': []}
        
        if has_header:
            header, *body = rows
            parsed = [list(header), *body]
        else:
            parsed = rows
        
        self.update_state(state='PROGRESS', meta={'progress': 1.0, 'message': 'Завершено'})
        
        return {'parsed': parsed}
        
    except Exception as e:
        logger.error(f"Ошибка обработки предпросмотра файла {temp_path}: {str(e)}")
        raise

